"""MH Program.xlsx 양식에 맞춘 엑셀 출력."""
from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from quality_mh.models import CalcResult, QualitativeRecord, QuantitativeRecord
from quality_mh.plant_config import PlantConfig
from quality_mh.summary_engine import SummaryReport, build_summary_report

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
TITLE_FONT = Font(bold=True, size=12)


def _style_header(ws, row: int, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def export_mh_program_excel(
    config: PlantConfig,
    calc_results: list[CalcResult],
    quantitative: list[QuantitativeRecord],
    qualitative: list[QualitativeRecord],
    report: SummaryReport | None = None,
    output_path: str | Path | None = None,
) -> BytesIO:
    """엑셀 MH Program 양식 5시트 출력."""
    report = report or build_summary_report(config, calc_results, qualitative)
    wb = Workbook()
    wb.remove(wb.active)

    # ── 종합 시트 ──
    ws = wb.create_sheet(f"종합({config.work_hours_per_day}hrs)")
    ws["B3"] = f"■ {config.plant_name} 품질 M/H 분석 ({config.analysis_year})"
    ws["B3"].font = TITLE_FONT
    ws["L3"] = f"{config.work_hours_per_day}/{config.work_hours_per_day}"

    headers = ["구분", "업무 항목", "", "현재원", "표준인원", "차이", "비고"]
    for c, h in enumerate(headers, 2):
        ws.cell(row=5, column=c, value=h)
    _style_header(ws, 5, len(headers) + 1)

    row = 6
    for qr in report.quantitative_rows:
        if qr.sub_label in ("입고", "공정", "완성", "시험"):
            ws.cell(row=row, column=4, value="표준")
            ws.cell(row=row, column=5, value="정 량")
            ws.cell(row=row, column=6, value=qr.sub_label)
            ws.cell(row=row, column=8, value=qr.current)
            ws.cell(row=row, column=9, value=round(qr.standard, 2))
            ws.cell(row=row, column=10, value=round(qr.diff, 2))
            ws.cell(row=row, column=11, value=qr.comment)
            row += 1
    if report.qualitative_row:
        q = report.qualitative_row
        ws.cell(row=row, column=5, value="정 성 合")
        ws.cell(row=row, column=8, value=q.current)
        ws.cell(row=row, column=9, value=q.standard)
        ws.cell(row=row, column=10, value=q.diff)
        row += 1
    if report.total_row:
        t = report.total_row
        ws.cell(row=row, column=5, value="총 계")
        ws.cell(row=row, column=8, value=t.current)
        ws.cell(row=row, column=9, value=round(t.standard, 2))
        ws.cell(row=row, column=10, value=round(t.diff, 2))

    # ── 정량 상세 ──
    ws_q = wb.create_sheet("정량 업무 MH 분석(상세)")
    q_headers = [
        "판정", "공장", "W/G", "라인", "업무 항목", "업무명", "산정기준", "발생빈도",
        "근거자료", "M/H산출식", "수행인원", "단위시간(분)", "Hr/회", "수행주기",
        "발생빈도(연)", "표준작업시간", "표준공수", "표준인원", "비고",
    ]
    for c, h in enumerate(q_headers, 2):
        ws_q.cell(row=4, column=c, value=h)
    _style_header(ws_q, 4, len(q_headers) + 1)

    calc_by_id = {r.record_id: r for r in calc_results}
    for i, rec in enumerate(quantitative, 10):
        calc = calc_by_id.get(rec.record_id)
        ws_q.cell(row=i, column=3, value=rec.plant)
        ws_q.cell(row=i, column=4, value=rec.wg)
        ws_q.cell(row=i, column=5, value=rec.line or "")
        ws_q.cell(row=i, column=6, value=rec.task_name)
        ws_q.cell(row=i, column=7, value=rec.sub_task or "")
        ws_q.cell(row=i, column=8, value=rec.estimation_method or "")
        ws_q.cell(row=i, column=9, value=rec.frequency_method_text or "")
        ws_q.cell(row=i, column=10, value=rec.data_source or "")
        ws_q.cell(row=i, column=11, value=rec.mh_formula or "")
        ws_q.cell(row=i, column=12, value=rec.performers)
        ws_q.cell(row=i, column=13, value=rec.unit_time_min)
        hr_per = rec.performers * rec.unit_time_min / 60
        ws_q.cell(row=i, column=14, value=round(hr_per, 4))
        ws_q.cell(row=i, column=15, value=rec.cycle_type or "")
        ws_q.cell(row=i, column=16, value=rec.annual_frequency or "")
        if calc:
            ws_q.cell(row=i, column=17, value=round(calc.standard_work_time_hr, 2))
            ws_q.cell(row=i, column=18, value=round(calc.standard_mh, 4))
            ws_q.cell(row=i, column=19, value=calc.standard_headcount)
        ws_q.cell(row=i, column=20, value=rec.hq_review or rec.remark or "")

    # ── 정성 상세 ──
    ws_ql = wb.create_sheet("정성 업무 MH 분석(상세)")
    ql_headers = ["공장", "W/G", "업무 항목", "업무 정의", "업무 내용", "표준인원(명)", "비고"]
    for c, h in enumerate(ql_headers, 2):
        ws_ql.cell(row=4, column=c, value=h)
    _style_header(ws_ql, 4, len(ql_headers) + 1)
    for i, rec in enumerate(qualitative, 6):
        ws_ql.cell(row=i, column=3, value=rec.plant)
        ws_ql.cell(row=i, column=4, value=rec.wg)
        ws_ql.cell(row=i, column=5, value=rec.task_name)
        ws_ql.cell(row=i, column=6, value=rec.task_definition or "")
        ws_ql.cell(row=i, column=7, value=rec.workload_desc or "")
        ws_ql.cell(row=i, column=8, value=rec.standard_headcount)
        ws_ql.cell(row=i, column=9, value=rec.remark or "")

    # ── 정수화 ──
    ws_r = wb.create_sheet("정수화")
    ws_r.append(["레코드ID", "업무", "W/G", "표준공수(M/H)", "정수화전", "표준인원", "정수화정책"])
    _style_header(ws_r, 1, 7)
    for r in calc_results:
        ws_r.append([
            r.record_id,
            r.frequency_factors_used.get("task_name", ""),
            r.frequency_factors_used.get("wg", ""),
            round(r.standard_mh, 4),
            round(r.standard_headcount_raw, 4),
            r.standard_headcount,
            config.effective_rounding_policy().value,
        ])

    # ── 그래프 데이터 ──
    ws_g = wb.create_sheet("그래프")
    ws_g.append(["구분", "현재원", "표준인원"])
    _style_header(ws_g, 1, 3)
    for qr in report.quantitative_rows:
        if qr.sub_label in ("입고", "공정", "완성", "시험"):
            ws_g.append([qr.sub_label, qr.current, round(qr.standard, 1)])
    if report.qualitative_row:
        ws_g.append(["정성", report.qualitative_row.current, report.qualitative_row.standard])
    ws_g.append([])
    ws_g.append(["Pareto 업무", "표준공수", "비중(%)"])
    for p in report.pareto:
        ws_g.append([p["업무"], p["표준공수"], p["비중(%)"]])

    for sheet in wb.worksheets:
        for col in sheet.columns:
            letter = get_column_letter(col[0].column)
            max_len = max(len(str(c.value or "")) for c in col)
            sheet.column_dimensions[letter].width = min(max_len + 2, 45)

    buf = BytesIO()
    if output_path:
        wb.save(output_path)
    wb.save(buf)
    buf.seek(0)
    return buf
