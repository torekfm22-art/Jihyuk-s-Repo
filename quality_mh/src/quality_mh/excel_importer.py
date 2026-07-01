"""엑셀 업로드 및 파싱."""
from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import BinaryIO

from openpyxl import load_workbook

from quality_mh.models import FrequencyDB, FrequencyMethod, QualitativeRecord, QuantitativeRecord

COLUMN_MAP: dict[str, dict[str, str]] = {
    "quantitative": {
        "record_id": "record_id",
        "레코드ID": "record_id",
        "공장": "plant",
        "plant": "plant",
        "W/G": "wg",
        "wg": "wg",
        "업무코드": "task_code",
        "task_code": "task_code",
        "업무항목": "task_name",
        "task_name": "task_name",
        "세부업무": "sub_task",
        "라인": "line",
        "line": "line",
        "라인그룹": "line_group",
        "단위시간(분)": "unit_time_min",
        "단위시간": "unit_time_min",
        "unit_time_min": "unit_time_min",
        "현재원": "current_headcount",
        "발생빈도_override": "frequency_override",
        "부가공수_override": "allowance_override",
        "비고": "remark",
    },
    "qualitative": {
        "record_id": "record_id",
        "공장": "plant",
        "W/G": "wg",
        "업무항목": "task_name",
        "업무정의": "task_definition",
        "업무량설명": "workload_desc",
        "기준인원": "standard_headcount",
        "현재인원": "current_headcount",
        "차이": "diff",
        "선정사유": "selection_reason",
        "비고": "remark",
    },
    "freq_db": {
        "task_code": "task_code",
        "업무코드": "task_code",
        "y1_actual": "y1_actual",
        "Y-1": "y1_actual",
        "y2_actual": "y2_actual",
        "Y-2": "y2_actual",
        "y3_actual": "y3_actual",
        "Y-3": "y3_actual",
        "ref_ratio": "ref_ratio",
        "기준비율": "ref_ratio",
        "plan_qty": "plan_qty",
        "계획생산량": "plan_qty",
        "cycle_type": "cycle_type",
        "수행주기": "cycle_type",
        "cycle_count": "cycle_count",
        "횟수": "cycle_count",
    },
}

SHEET_ALIASES: dict[str, str] = {
    "정량_상세": "quantitative",
    "정량상세": "quantitative",
    "정량": "quantitative",
    "정성_상세": "qualitative",
    "정성상세": "qualitative",
    "정성": "qualitative",
    "기준표": "freq_db",
    "발생빈도DB": "freq_db",
    "freq_db": "freq_db",
}


def _normalize_header(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _map_row(headers: list[str], row: tuple, mapping: dict[str, str]) -> dict:
    data: dict = {}
    for idx, header in enumerate(headers):
        key = mapping.get(header)
        if key and idx < len(row):
            val = row[idx]
            if val is not None and str(val).strip() != "":
                data[key] = val
    return data


def _coerce_float(val) -> float | None:
    if val is None or str(val).strip() == "":
        return None
    return float(val)


def _coerce_int(val) -> int | None:
    if val is None or str(val).strip() == "":
        return None
    return int(float(val))


def _parse_quantitative_row(data: dict, row_idx: int) -> QuantitativeRecord | None:
    if not data.get("plant") and not data.get("task_code"):
        return None
    return QuantitativeRecord(
        record_id=str(data.get("record_id") or f"IMP-Q-{row_idx}"),
        plant=str(data.get("plant", "")),
        wg=str(data.get("wg", "")),
        task_code=str(data.get("task_code", "")),
        task_name=str(data.get("task_name", "")),
        sub_task=str(data["sub_task"]) if data.get("sub_task") else None,
        line=str(data["line"]) if data.get("line") else None,
        line_group=str(data["line_group"]) if data.get("line_group") else None,
        unit_time_min=float(data.get("unit_time_min", 0) or 0),
        current_headcount=_coerce_float(data.get("current_headcount")) or 0.0,
        frequency_override=_coerce_float(data.get("frequency_override")),
        allowance_override=_coerce_float(data.get("allowance_override")),
        remark=str(data["remark"]) if data.get("remark") else None,
    )


def _parse_qualitative_row(data: dict, row_idx: int) -> QualitativeRecord | None:
    if not data.get("plant") and not data.get("task_name"):
        return None
    return QualitativeRecord(
        record_id=str(data.get("record_id") or f"IMP-QL-{row_idx}"),
        plant=str(data.get("plant", "")),
        wg=str(data.get("wg", "")),
        task_name=str(data.get("task_name", "")),
        task_definition=str(data["task_definition"]) if data.get("task_definition") else None,
        workload_desc=str(data["workload_desc"]) if data.get("workload_desc") else None,
        standard_headcount=_coerce_int(data.get("standard_headcount")) or 0,
        current_headcount=_coerce_int(data.get("current_headcount")) or 0,
        diff=_coerce_int(data.get("diff")) or 0,
        selection_reason=str(data["selection_reason"]) if data.get("selection_reason") else None,
        remark=str(data["remark"]) if data.get("remark") else None,
    )


def _parse_freq_db_row(data: dict) -> FrequencyDB | None:
    if not data.get("task_code"):
        return None
    method_str = str(data.get("frequency_method", FrequencyMethod.WEIGHTED_AVG.value))
    try:
        method = FrequencyMethod(method_str)
    except ValueError:
        method = FrequencyMethod.WEIGHTED_AVG
    return FrequencyDB(
        task_code=str(data["task_code"]),
        frequency_method=method,
        y1_actual=_coerce_float(data.get("y1_actual")),
        y2_actual=_coerce_float(data.get("y2_actual")),
        y3_actual=_coerce_float(data.get("y3_actual")),
        ref_ratio=_coerce_float(data.get("ref_ratio")),
        plan_qty=_coerce_float(data.get("plan_qty")),
        cycle_type=str(data["cycle_type"]) if data.get("cycle_type") else None,
        cycle_count=_coerce_float(data.get("cycle_count")),
        data_source=str(data["data_source"]) if data.get("data_source") else None,
        description=str(data["description"]) if data.get("description") else None,
    )


def import_excel(
    source: str | Path | BinaryIO | BytesIO,
) -> tuple[list[QuantitativeRecord], list[QualitativeRecord], list[FrequencyDB]]:
    """엑셀 파일에서 정량/정성/발생빈도 DB 데이터를 파싱."""
    wb = load_workbook(source, read_only=True, data_only=True)
    quantitative: list[QuantitativeRecord] = []
    qualitative: list[QualitativeRecord] = []
    freq_db: list[FrequencyDB] = []

    for sheet_name in wb.sheetnames:
        sheet_type = SHEET_ALIASES.get(sheet_name.strip())
        if sheet_type is None:
            for alias, stype in SHEET_ALIASES.items():
                if alias in sheet_name:
                    sheet_type = stype
                    break
        if sheet_type is None:
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        headers = [_normalize_header(h) for h in rows[0]]
        mapping = COLUMN_MAP[sheet_type]

        for row_idx, row in enumerate(rows[1:], start=2):
            data = _map_row(headers, row, mapping)
            if sheet_type == "quantitative":
                rec = _parse_quantitative_row(data, row_idx)
                if rec:
                    quantitative.append(rec)
            elif sheet_type == "qualitative":
                rec = _parse_qualitative_row(data, row_idx)
                if rec:
                    qualitative.append(rec)
            elif sheet_type == "freq_db":
                entry = _parse_freq_db_row(data)
                if entry:
                    freq_db.append(entry)

    wb.close()
    return quantitative, qualitative, freq_db
