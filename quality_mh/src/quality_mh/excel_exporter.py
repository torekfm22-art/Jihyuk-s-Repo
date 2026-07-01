"""결과 엑셀 생성."""
from __future__ import annotations

from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from quality_mh.calculation_engine import (
    aggregate_by_line,
    aggregate_by_plant,
    aggregate_by_wg,
)
from quality_mh.display_labels import COLUMN_LABELS_KO
from quality_mh.models import CalcResult, FrequencyDB, QualitativeRecord, RuleMaster
from quality_mh.rule_master import build_rule_master

HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
RED_FONT = Font(color="FF0000")


def _ko(key: str) -> str:
    return COLUMN_LABELS_KO.get(key, key)


def _style_header(ws, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _write_rows(ws, headers: list[str], rows: list[list]) -> None:
    ws.append(headers)
    _style_header(ws, len(headers))
    for row in rows:
        ws.append(row)
        excel_row = ws.max_row
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=excel_row, column=col_idx)
            if isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal="right")
                if isinstance(val, float) and val < 0:
                    cell.font = RED_FONT


def _autosize(ws) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[letter].width = min(max_len + 2, 40)


def export_excel(
    calc_results: list[CalcResult],
    qualitative_records: list[QualitativeRecord] | None = None,
    rules: list[RuleMaster] | None = None,
    freq_db: list[FrequencyDB] | None = None,
    history: list[dict] | None = None,
    output_path: str | Path | None = None,
) -> BytesIO:
    """7개 시트 구성의 결과 엑셀 생성."""
    wb = Workbook()
    wb.remove(wb.active)

    qualitative_records = qualitative_records or []
    rules = rules or build_rule_master()
    freq_db = freq_db or []
    history = history or []

    ws_sum = wb.create_sheet("종합")
    plant_agg = aggregate_by_plant(calc_results)
    _write_rows(
        ws_sum,
        ["공장", "구분", "표준공수(M/H)", "표준인원", "건수"],
        [
            [k, "정량", v["standard_mh"], v["standard_headcount"], v["record_count"]]
            for k, v in plant_agg.items()
        ]
        + [
            [r.plant, "정성", 0, r.standard_headcount, 1]
            for r in qualitative_records
        ],
    )
    _autosize(ws_sum)

    ws_q = wb.create_sheet("정량_상세")
    q_headers = [
        _ko("record_id"), _ko("auto_frequency"), _ko("final_frequency"), _ko("is_overridden"),
        _ko("unit_time_hr"), _ko("standard_work_time_hr"), _ko("allowance_rate"),
        _ko("final_work_time_hr"), _ko("standard_mh"), _ko("standard_md"),
        _ko("standard_headcount"), _ko("diff_from_current"), _ko("calc_log"),
    ]
    _write_rows(
        ws_q,
        q_headers,
        [
            [
                r.record_id, r.auto_frequency, r.final_frequency,
                "예" if r.is_overridden else "아니오",
                r.unit_time_hr, r.standard_work_time_hr, r.allowance_rate,
                r.final_work_time_hr, r.standard_mh, r.standard_md,
                r.standard_headcount, r.diff_from_current,
                " | ".join(r.calc_log),
            ]
            for r in calc_results
        ],
    )
    _autosize(ws_q)

    ws_ql = wb.create_sheet("정성_상세")
    ql_headers = [
        _ko("record_id"), _ko("plant"), _ko("wg"), _ko("task_name"),
        _ko("standard_headcount"), _ko("current_headcount"), _ko("diff"),
        _ko("selection_reason"), _ko("remark"),
    ]
    _write_rows(
        ws_ql,
        ql_headers,
        [
            [
                r.record_id, r.plant, r.wg, r.task_name, r.standard_headcount,
                r.current_headcount, r.diff, r.selection_reason, r.remark,
            ]
            for r in qualitative_records
        ],
    )
    _autosize(ws_ql)

    ws_line = wb.create_sheet("라인별_집계")
    line_agg = aggregate_by_line(calc_results)
    _write_rows(
        ws_line,
        [_ko("line"), _ko("plant"), _ko("wg"), _ko("standard_mh"),
         _ko("standard_headcount"), _ko("record_count")],
        [
            [v["line"], v["plant"], v["wg"], v["standard_mh"],
             v["standard_headcount"], v["record_count"]]
            for v in line_agg.values()
        ],
    )
    _autosize(ws_line)

    ws_wg = wb.create_sheet("WG별_집계")
    wg_agg = aggregate_by_wg(calc_results)
    _write_rows(
        ws_wg,
        [_ko("wg"), _ko("plant"), _ko("standard_mh"),
         _ko("standard_headcount"), _ko("record_count")],
        [
            [v["wg"], v["plant"], v["standard_mh"],
             v["standard_headcount"], v["record_count"]]
            for v in wg_agg.values()
        ],
    )
    _autosize(ws_wg)

    ws_diff = wb.create_sheet("차이분석")
    _write_rows(
        ws_diff,
        [_ko("record_id"), _ko("plant"), _ko("standard_headcount"),
         _ko("current_headcount"), _ko("diff")],
        [
            [
                r.record_id,
                r.frequency_factors_used.get("plant", ""),
                r.standard_headcount,
                r.frequency_factors_used.get("current_headcount", 0),
                r.diff_from_current,
            ]
            for r in calc_results
        ],
    )
    _autosize(ws_diff)

    ws_ref = wb.create_sheet("기준표")
    _write_rows(
        ws_ref,
        [_ko("task_code"), _ko("wg"), _ko("task_name"), _ko("frequency_method"),
         _ko("default_allowance_rate")],
        [
            [r.task_code, r.wg, r.task_name, r.frequency_method.value, r.default_allowance_rate]
            for r in rules
        ],
    )
    row_offset = len(rules) + 3
    ws_ref.cell(row=row_offset, column=1, value="--- 발생빈도 DB ---")
    freq_headers = [
        _ko("task_code"), _ko("frequency_method"), _ko("y1"), _ko("y2"), _ko("y3"),
        _ko("ref_ratio"), _ko("plan_qty"), _ko("cycle_type"), _ko("cycle_count"),
    ]
    for col, h in enumerate(freq_headers, 1):
        ws_ref.cell(row=row_offset + 1, column=col, value=h)
    for f in freq_db:
        ws_ref.append([
            f.task_code, f.frequency_method.value,
            f.y1_actual, f.y2_actual, f.y3_actual,
            f.ref_ratio, f.plan_qty, f.cycle_type, f.cycle_count,
        ])
    _autosize(ws_ref)

    ws_hist = wb.create_sheet("검토이력")
    _write_rows(
        ws_hist,
        [_ko("history_id"), _ko("record_id"), _ko("field_name"), _ko("old_value"),
         _ko("new_value"), _ko("changed_at"), _ko("change_reason")],
        [
            [h.get("history_id"), h.get("record_id"), h.get("field_name"),
             h.get("old_value"), h.get("new_value"), h.get("changed_at"), h.get("change_reason")]
            for h in history
        ],
    )
    _autosize(ws_hist)

    buf = BytesIO()
    if output_path:
        wb.save(output_path)
    wb.save(buf)
    buf.seek(0)
    return buf
