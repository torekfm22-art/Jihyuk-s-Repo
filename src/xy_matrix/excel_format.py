"""
Excel 보고서 표 서식 (openpyxl).
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(name="Malgun Gothic", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Malgun Gothic", size=10)
HIGHLIGHT_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
ALT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="B4B4B4"),
    right=Side(style="thin", color="B4B4B4"),
    top=Side(style="thin", color="B4B4B4"),
    bottom=Side(style="thin", color="B4B4B4"),
)


def _auto_column_width(ws: Worksheet, max_width: int = 42) -> None:
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        length = 0
        for cell in col_cells:
            val = cell.value
            if val is not None:
                length = max(length, len(str(val)))
        ws.column_dimensions[letter].width = min(max(length + 2, 10), max_width)


def style_data_table(
    ws: Worksheet,
    *,
    header_row: int = 1,
    highlight_col: int | None = None,
    highlight_values: set[Any] | None = None,
    number_cols: dict[int, str] | None = None,
) -> None:
    """헤더·테두리·교차 행·선택 열 강조."""
    if ws.max_row < 1:
        return
    max_col = ws.max_column

    for col in range(1, max_col + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    for row in range(header_row + 1, ws.max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if (row - header_row) % 2 == 0:
                cell.fill = ALT_FILL
            if highlight_col and col == highlight_col:
                if highlight_values is None or cell.value in highlight_values:
                    cell.fill = HIGHLIGHT_FILL
            if number_cols and col in number_cols:
                cell.number_format = number_cols[col]

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    _auto_column_width(ws)


def embed_image_below_table(
    ws: Worksheet,
    image_path: str | Path,
    *,
    gap_rows: int = 2,
    anchor_row: int | None = None,
    anchor_col: str = "A",
    max_width_px: int = 640,
) -> None:
    path = Path(image_path)
    if not path.exists():
        return
    if anchor_row is None:
        anchor_row = ws.max_row + gap_rows + 1
    else:
        anchor_row = anchor_row + gap_rows
    img = XLImage(str(path))
    scale = min(1.0, max_width_px / max(img.width, 1))
    img.width = int(img.width * scale)
    img.height = int(img.height * scale)
    ws.add_image(img, f"{anchor_col}{anchor_row}")
