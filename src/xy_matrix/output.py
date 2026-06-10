"""
분석 결과 Excel·표 형식 출력.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from src.xy_matrix.excel_format import embed_image_below_table, style_data_table
from src.xy_matrix.excel_verification import (
    VERIFY_SHEET,
    append_raw_data_sheet,
    write_verification_sheet,
)
from src.xy_matrix.report_summary import (
    build_key_findings_rows,
    ctp_recommendations_df,
    kv_to_dataframe,
    monitoring_df,
    multi_reg_summary_df,
)

SECTION_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
KV_FILL = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
FINDINGS_TITLE_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
FINDINGS_SECTION_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")


@dataclass
class SheetLayout:
    matrix_header_row: int = 1
    table_header_rows: list[int] = field(default_factory=list)
    section_title_rows: list[int] = field(default_factory=list)
    key_findings_start_row: int | None = None
    key_findings_row_count: int = 0


def matrix_to_display_df(matrix_df: pd.DataFrame) -> pd.DataFrame:
    rename = {
        "rank": "순위",
        "x_type": "인자 유형",
        "x_column": "X 인자명",
        "method": "분석기법",
        "r_square": "R-Square",
        "p_value": "P-Value",
        "score": "점수",
        "symbol": "기호",
        "interpretation": "해석",
    }
    cols = [c for c in rename if c in matrix_df.columns]
    out = matrix_df[cols].rename(columns=rename)
    if "P-Value" in out.columns:
        out["P-Value"] = out["P-Value"].apply(_format_p_value)
    if "R-Square" in out.columns:
        out["R-Square"] = out["R-Square"].apply(
            lambda v: round(float(v), 4) if pd.notna(v) else None
        )
    return out


def _format_p_value(v: Any) -> str | float | None:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    try:
        fv = float(v)
        if fv < 1e-10:
            return "≈ 0"
        return fv
    except (TypeError, ValueError):
        return str(v)


def pareto_to_display_df(pareto_df: pd.DataFrame) -> pd.DataFrame:
    x_col = "X 인자명" if "X 인자명" in pareto_df.columns else pareto_df.columns[0]
    out = pareto_df.copy()
    if "순위" not in out.columns:
        out.insert(0, "순위", range(1, len(out) + 1))
    out = out.rename(columns={x_col: "X 인자명"})
    out["누적 기여(%)"] = out["cumulative_pct"].apply(
        lambda v: round(float(v), 1) if pd.notna(v) else 0.0
    )
    if "pareto_80" in out.columns:
        out["80% 구간"] = out["pareto_80"].map(lambda x: "●" if x else "")
    cols = ["순위", "X 인자명", "score", "누적 기여(%)", "80% 구간"]
    cols = [c for c in cols if c in out.columns]
    return out[cols].rename(columns={"score": "점수"})


def build_pareto_data(matrix_df: pd.DataFrame) -> pd.DataFrame:
    x_key = "x_column" if "x_column" in matrix_df.columns else matrix_df.columns[0]
    pareto = matrix_df[[x_key, "score"]].copy()
    pareto = pareto.sort_values("score", ascending=False).reset_index(drop=True)
    pareto.insert(0, "순위", range(1, len(pareto) + 1))
    total = pareto["score"].sum()
    if total <= 0:
        pareto["cumulative_pct"] = 0.0
        pareto["pareto_80"] = False
        return pareto.rename(columns={x_key: "X 인자명"})

    pareto["cumulative_pct"] = pareto["score"].cumsum() / total * 100
    pareto["pareto_80"] = pareto["cumulative_pct"] <= 80
    if pareto["pareto_80"].any():
        last_80 = pareto[pareto["pareto_80"]].index[-1]
        pareto.loc[:last_80, "pareto_80"] = True
    return pareto.rename(columns={x_key: "X 인자명"})


def _section_title_df(title: str) -> pd.DataFrame:
    return pd.DataFrame([[title, ""]], columns=["항목", "내용"])


def _write_xy_consolidated_sheet(
    result: dict[str, Any], writer: pd.ExcelWriter
) -> SheetLayout:
    """XY_매트릭스 한 시트에 표·파레토·요약·CTP·모니터링·다중회귀 통합."""
    sheet = "XY_매트릭스"
    layout = SheetLayout(matrix_header_row=1)
    row = 0

    display = matrix_to_display_df(result["matrix"])
    display.to_excel(writer, sheet_name=sheet, index=False, startrow=row)
    row += len(display) + 2

    title_pareto = _section_title_df("【 파레토 분석 】")
    title_pareto.to_excel(writer, sheet_name=sheet, index=False, header=False, startrow=row)
    layout.section_title_rows.append(row + 1)
    row += 2

    pareto_disp = pareto_to_display_df(result["pareto_data"])
    pareto_disp.to_excel(writer, sheet_name=sheet, index=False, startrow=row)
    layout.table_header_rows.append(row + 1)
    row += len(pareto_disp) + 2

    mr_df = multi_reg_summary_df(result)
    if mr_df is not None:
        title_mr = _section_title_df("【 다중회귀 요약 】")
        title_mr.to_excel(writer, sheet_name=sheet, index=False, header=False, startrow=row)
        layout.section_title_rows.append(row + 1)
        row += 2
        mr_df.to_excel(writer, sheet_name=sheet, index=False, startrow=row)
        layout.table_header_rows.append(row + 1)
        row += len(mr_df) + 2

    ctp_df = ctp_recommendations_df(result)
    if ctp_df is not None:
        title_ctp = _section_title_df("【 CTP · SPC 권고 】")
        title_ctp.to_excel(writer, sheet_name=sheet, index=False, header=False, startrow=row)
        layout.section_title_rows.append(row + 1)
        row += 2
        ctp_df.to_excel(writer, sheet_name=sheet, index=False, startrow=row)
        layout.table_header_rows.append(row + 1)
        row += len(ctp_df) + 2

    mon_df = monitoring_df(result)
    if mon_df is not None:
        title_mon = _section_title_df("【 모니터링 (제어 불가) 】")
        title_mon.to_excel(writer, sheet_name=sheet, index=False, header=False, startrow=row)
        layout.section_title_rows.append(row + 1)
        row += 2
        mon_df.to_excel(writer, sheet_name=sheet, index=False, startrow=row)
        layout.table_header_rows.append(row + 1)
        row += len(mon_df) + 2

    row += 1
    findings = kv_to_dataframe(build_key_findings_rows(result))
    layout.key_findings_start_row = row + 1
    layout.key_findings_row_count = len(findings)
    findings.to_excel(writer, sheet_name=sheet, index=False, header=False, startrow=row)

    return layout


def _write_sheets(result: dict[str, Any], writer: pd.ExcelWriter) -> SheetLayout:
    """XY_매트릭스(통합) + 수식검증만 별도 시트."""
    return _write_xy_consolidated_sheet(result, writer)


def _style_section_titles(ws, rows: list[int]) -> None:
    for r in rows:
        cell = ws.cell(r, 1)
        cell.font = Font(name="Malgun Gothic", bold=True, size=11)
        cell.fill = SECTION_FILL
        ws.merge_cells(f"A{r}:B{r}")
        ws.row_dimensions[r].height = 22


def _style_key_findings_block(ws, start_row: int, n_rows: int) -> None:
    """하단 핵심 발견 서술 블록 서식."""
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 72
    for offset in range(n_rows):
        row = start_row + offset
        label = ws.cell(row, 1).value
        content = ws.cell(row, 2).value
        is_title = offset == 0 and label and "핵심 발견" in str(label)
        is_section = bool(label) and str(label).strip() and not is_title

        if is_title:
            ws.merge_cells(f"A{row}:B{row}")
            cell = ws.cell(row, 1)
            cell.font = Font(name="Malgun Gothic", bold=True, size=12)
            cell.fill = FINDINGS_TITLE_FILL
            cell.alignment = Alignment(vertical="center")
            ws.row_dimensions[row].height = 28
            continue

        if not label and content:
            cell = ws.cell(row, 1)
            cell.value = content
            cell.font = Font(
                name="Malgun Gothic",
                size=10,
                bold=str(content).startswith(("☑", "💡", "📌", "△", "※")),
            )
            cell.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
            cell.fill = KV_FILL
            ws.merge_cells(f"A{row}:B{row}")
            h = 18 + min(80, max(0, len(str(content)) // 60) * 14)
            ws.row_dimensions[row].height = max(18, h)
            continue

        for col in (1, 2):
            c = ws.cell(row, col)
            c.fill = FINDINGS_SECTION_FILL if is_section else KV_FILL
            c.font = Font(
                name="Malgun Gothic",
                size=10,
                bold=col == 1 and is_section,
            )
            c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 24 if content or label else 8


def _apply_workbook_styles(
    path: Path, result: dict[str, Any], layout: SheetLayout
) -> None:
    wb = load_workbook(path)

    if "XY_매트릭스" in wb.sheetnames:
        ws = wb["XY_매트릭스"]
        sym_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(layout.matrix_header_row, col).value == "기호":
                sym_col = col
        style_data_table(
            ws,
            header_row=layout.matrix_header_row,
            highlight_col=sym_col,
            highlight_values={"◎", "○", "△"},
        )
        _style_section_titles(ws, layout.section_title_rows)

        if layout.key_findings_start_row and layout.key_findings_row_count:
            _style_key_findings_block(
                ws,
                layout.key_findings_start_row,
                layout.key_findings_row_count,
            )

        for hdr_row in layout.table_header_rows:
            flag_col = None
            for col in range(1, ws.max_column + 1):
                if ws.cell(hdr_row, col).value == "80% 구간":
                    flag_col = col
            style_data_table(
                ws,
                header_row=hdr_row,
                highlight_col=flag_col,
                highlight_values={"●"} if flag_col else None,
            )

        chart = result.get("pareto_chart_path")
        if chart and layout.key_findings_start_row:
            embed_image_below_table(
                ws,
                chart,
                gap_rows=0,
                anchor_row=layout.key_findings_start_row,
                anchor_col="I",
                max_width_px=400,
            )

    write_verification_sheet(
        wb, result, matrix_data_start_row=layout.matrix_header_row
    )

    analysis_df = result.get("analysis_data")
    if analysis_df is not None:
        append_raw_data_sheet(wb, analysis_df)

    if VERIFY_SHEET in wb.sheetnames:
        style_data_table(wb[VERIFY_SHEET])

    if "Raw_data" in wb.sheetnames:
        style_data_table(wb["Raw_data"])

    wb.save(path)


def export_to_excel(result: dict[str, Any], output_path: str | Path) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        layout = _write_sheets(result, writer)

    _apply_workbook_styles(path, result, layout)
    return path


def format_matrix_as_text(matrix_display: pd.DataFrame, max_rows: int = 20) -> str:
    df = matrix_display.head(max_rows)
    cols = list(df.columns)
    widths = {c: max(len(str(c)), *(len(str(v)) for v in df[c])) for c in cols}
    widths = {c: min(w, 24) for c, w in widths.items()}
    sep = "+" + "+".join("-" * (widths[c] + 2) for c in cols) + "+"
    header = "|" + "|".join(f" {str(c)[:widths[c]]:^{widths[c]}} " for c in cols) + "|"
    lines = [sep, header, sep]
    for _, row in df.iterrows():
        lines.append(
            "|" + "|".join(
                f" {str(row[c])[:widths[c]]:^{widths[c]}} " for c in cols
            ) + "|"
        )
    lines.append(sep)
    return "\n".join(lines)
