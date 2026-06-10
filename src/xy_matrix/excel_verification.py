"""
Excel 수식으로 프로그램 산출값 검증 시트 생성.
"""
from __future__ import annotations

from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

from src.xy_matrix.constants import P_VALUE_ALPHA, TYPE_CONTINUOUS
from src.xy_matrix.excel_format import style_data_table

RAW_SHEET = "Raw_data"
MATRIX_SHEET = "XY_매트릭스"
VERIFY_SHEET = "수식검증"
P_TOL = 0.0001
R2_TOL = 0.0001


def _col_letter(df: pd.DataFrame, col_name: str) -> str | None:
    if col_name not in df.columns:
        return None
    return get_column_letter(df.columns.get_loc(col_name) + 1)


def _data_range(sheet: str, letter: str, last_row: int) -> str:
    return f"'{sheet}'!${letter}$2:${letter}${last_row}"


def _excel_corr_p_formula(y_range: str, x_range: str, n_expr: str) -> str:
    """
    상관 기반 양측 P-value. |r|≈1 또는 분모≈0 이면 0 반환 (문자 N/A 없음).
    """
    r = f"CORREL({y_range},{x_range})"
    den = f"(1-{r}^2)"
    t_stat = f"IF(OR({n_expr}<3,{den}<=1E-12),0,ABS({r})*SQRT(({n_expr}-2)/{den}))"
    p_val = (
        f"IF(OR({n_expr}<3,ABS({r})>=1-1E-12,{den}<=1E-12),0,"
        f"IFERROR(T.DIST.2T({t_stat},{n_expr}-2),0))"
    )
    return f"={p_val}"


def _p_verify_formula(row: int) -> str:
    h, g = f"H{row}", f"G{row}"
    return (
        f'=IF(OR(NOT(ISNUMBER({h})),{h}=""),"—",'
        f'IF(ABS({g}-{h})<{P_TOL},"OK","NG"))'
    )


def _score_excel_formula(row: int) -> str:
    """Excel P 유효 시 H, 아니면 프로그램 P(G)로 1-3-9 점수 산출."""
    e, g, h = f"E{row}", f"G{row}", f"H{row}"
    a = P_VALUE_ALPHA

    def _score_body(p_cell: str) -> str:
        return (
            f"IF(OR(NOT(ISNUMBER({p_cell})),{p_cell}>={a}),0,"
            f"IF({e}>=0.7,9,IF({e}>=0.4,3,1)))"
        )

    return (
        f"=IF(OR(NOT(ISNUMBER({h})),{h}=\"\"),"
        f"{_score_body(g)},"
        f"{_score_body(h)})"
    )


def _r2_verify_formula(row: int) -> str:
    return f'=IF(ABS(D{row}-E{row})<{R2_TOL},"OK","NG")'


def write_verification_sheet(
    wb: Workbook,
    result: dict[str, Any],
    *,
    matrix_data_start_row: int,
) -> None:
    """수식검증 시트: 프로그램값 vs Excel 수식값 대조."""
    if VERIFY_SHEET in wb.sheetnames:
        del wb[VERIFY_SHEET]
    ws = wb.create_sheet(VERIFY_SHEET)

    df: pd.DataFrame | None = result.get("analysis_data")
    y_col = result.get("y_column", "")
    matrix = result.get("matrix")
    if df is None or matrix is None or not len(matrix):
        ws["A1"] = "검증 불가: 분석 데이터 없음"
        return

    last_row = len(df) + 1
    y_letter = _col_letter(df, y_col)
    n_expr = f"COUNTA({_data_range(RAW_SHEET, y_letter or 'A', last_row)})"

    ws["A1"] = "【 수식 검증 안내 】"
    ws["A1"].font = Font(name="Malgun Gothic", bold=True, size=11)
    ws.merge_cells("A1:M1")
    ws["A2"] = (
        "계량×계량: RSQ·상관 P·1-3-9 점수를 Excel로 재계산합니다. "
        f"R²·P 허용오차 ±{R2_TOL}, 점수는 정수 일치. 범주형(ANOVA 등)은 — 표시."
    )
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:M2")

    headers = [
        "X 인자", "유형", "분석기법", "프로그램 R²", "Excel R²", "R² 검증",
        "프로그램 P", "Excel P", "P 검증", "프로그램 점수", "Excel 점수", "점수 검증",
        "비고",
    ]
    start = 4
    for c, h in enumerate(headers, 1):
        ws.cell(row=start, column=c, value=h)

    m_header = matrix_data_start_row
    for row_idx, (_, row_data) in enumerate(matrix.iterrows()):
        r = start + 1 + row_idx
        x_col = str(row_data["x_column"])
        x_type = str(row_data.get("x_type", ""))
        method = str(row_data.get("method_code", row_data.get("method", "")))

        ws.cell(r, 1, x_col)
        ws.cell(r, 2, x_type)
        ws.cell(r, 3, row_data.get("method", method))
        ws.cell(r, 4, row_data.get("r_square"))
        ws.cell(r, 7, row_data.get("p_value"))
        ws.cell(r, 10, row_data.get("score"))

        x_letter = _col_letter(df, x_col)
        note = ""

        if (
            y_letter
            and x_letter
            and x_type == TYPE_CONTINUOUS
            and method in ("linear_regression", "선형회귀분석")
        ):
            yr = _data_range(RAW_SHEET, y_letter, last_row)
            xr = _data_range(RAW_SHEET, x_letter, last_row)
            ws.cell(r, 5, f"=RSQ({yr},{xr})")
            ws.cell(r, 6, _r2_verify_formula(r))
            ws.cell(r, 8, _excel_corr_p_formula(yr, xr, n_expr))
            ws.cell(r, 9, _p_verify_formula(r))
            ws.cell(r, 11, _score_excel_formula(r))
            ws.cell(r, 12, f'=IF(J{r}=K{r},"OK","NG")')
        else:
            ws.cell(r, 5, "—")
            ws.cell(r, 6, "—")
            ws.cell(r, 8, "—")
            ws.cell(r, 9, "—")
            m_r = m_header + 1 + row_idx
            ws.cell(r, 11, (
                f"=IF(AND(ISNUMBER(G{r}),G{r}<{P_VALUE_ALPHA}),"
                f"IF(D{r}>=0.7,9,IF(D{r}>=0.4,3,1)),0)"
            ))
            ws.cell(r, 12, f'=IF(J{r}=K{r},"OK","NG")')
            note = "ANOVA/카이제곱 — 매트릭스 시트 값 참조"
            ws.cell(r, 4, f"='{MATRIX_SHEET}'!E{m_r}")
            ws.cell(r, 7, f"='{MATRIX_SHEET}'!F{m_r}")

        ws.cell(r, 13, note)

    style_data_table(ws, header_row=start)

    p_start = start + len(matrix) + 4
    ws.cell(p_start, 1, "【 파레토 누적% 검증 】").font = Font(
        name="Malgun Gothic", bold=True, size=11
    )
    pareto = result.get("pareto_data")
    if pareto is not None and len(pareto):
        ph = p_start + 1
        for c, h in enumerate(
            ["순위", "X 인자", "프로그램 점수", "프로그램 누적%", "Excel 누적%", "검증"], 1
        ):
            ws.cell(ph, c, h)
        total_row = ph + len(pareto) + 1
        for idx, (_, prow) in enumerate(pareto.iterrows()):
            pr = ph + 1 + idx
            ws.cell(pr, 1, int(prow.get("순위", idx + 1)))
            ws.cell(pr, 2, prow.get("X 인자명", ""))
            ws.cell(pr, 3, float(prow["score"]))
            ws.cell(pr, 4, float(prow.get("cumulative_pct", 0)))
            if idx == 0:
                ws.cell(pr, 5, f"=IF(C{total_row}=0,0,C{pr}/C${total_row}*100)")
            else:
                ws.cell(pr, 5, f"=E{pr-1}+IF(C${total_row}=0,0,C{pr}/C${total_row}*100)")
            ws.cell(pr, 6, f'=IF(ABS(D{pr}-E{pr})<0.2,"OK","NG")')
        ws.cell(total_row, 3, f"=SUM(C{ph+1}:C{ph+len(pareto)})")
        style_data_table(ws, header_row=ph)

    _auto_width(ws)


def _auto_width(ws: Worksheet, max_w: int = 48) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        w = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[letter].width = min(max(w + 2, 10), max_w)


def append_raw_data_sheet(wb: Workbook, df: pd.DataFrame, max_rows: int = 100_000) -> int:
    """Raw_data 시트 추가."""
    if RAW_SHEET in wb.sheetnames:
        del wb[RAW_SHEET]
    ws = wb.create_sheet(RAW_SHEET)

    truncated = len(df) > max_rows
    out = df.head(max_rows) if truncated else df

    for r_idx, row in enumerate(dataframe_to_rows(out, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    if truncated:
        note_row = len(out) + 3
        ws.cell(
            note_row,
            1,
            f"※ 상위 {max_rows:,}행만 표시 (전체 {len(df):,}행).",
        )

    style_data_table(ws, header_row=1)
    return len(out) + 1
