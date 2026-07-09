"""AIAG/VDA 스타일 종합 보고서 (Excel 종합시트 + PDF) 및 세부 시트."""
from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.spc.chart_expert_review import format_conclusions_for_report
from src.spc.decision_models import SpcDecisionResult
from src.spc.minitab_charts import ChartPaths
from src.spc.report_decision_dashboard import (
    add_decision_dashboard_sheet,
    apply_decision_sheet_styles,
    write_verdict_strip_on_summary,
)
from src.spc.report_glossary_sheet import add_glossary_sheet
from src.spc.report_validation_sheet import SheetMeta, add_validation_sheet, sanitize_xlsx_formula_file
from src.spc.statistics import SpcAnalysisResult, _cap_round
from src.spc.characteristic_split import safe_filename_slug
from src.spc.traceability_export import build_traceability_sheets

logger = logging.getLogger(__name__)

TRACE_CAUTION_FILL = PatternFill("solid", fgColor="FFE4E1")
TRACE_HEADER_FILL = PatternFill("solid", fgColor="9C0006")

THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
LABEL_FONT = Font(bold=True, size=9)
VALUE_FONT = Font(size=9)


def _fmt_spec_limit(val: float | None) -> str:
    if val is None:
        return "—"
    return f"{val:g}"


def _fmt_cap_number(val: float) -> str:
    rounded = _cap_round(val)
    if isinstance(rounded, str):
        return rounded
    return f"{rounded:.4f}"


def _norm_pct_outside_spec(
    cap,
    sp_stats,
) -> tuple[float | None, float | None]:
    """정규 근사 P% — 해당 규격한이 없으면 None."""
    if not cap.std_within or cap.std_within <= 0:
        return None, None
    scale = cap.std_within
    pct_above = (
        float(sp_stats.norm.sf((cap.usl - cap.mean) / scale) * 100)
        if cap.usl is not None
        else None
    )
    pct_below = (
        float(sp_stats.norm.cdf((cap.lsl - cap.mean) / scale) * 100)
        if cap.lsl is not None
        else None
    )
    return pct_above, pct_below


def _raw_values_from_sample(raw_sample: pd.DataFrame | None):
    if raw_sample is None or "value" not in raw_sample.columns:
        return None
    return raw_sample["value"].to_numpy()


def _conclusion_text(
    result: SpcAnalysisResult,
    raw_sample: pd.DataFrame | None = None,
    decision: SpcDecisionResult | None = None,
) -> str:
    base = format_conclusions_for_report(result, _raw_values_from_sample(raw_sample))
    if decision is None:
        return base
    return base + "\n\n" + format_decision_for_report(decision)


def format_decision_for_report(decision: SpcDecisionResult) -> str:
    """회사 기준 판정·해석 코멘트 보고서 텍스트."""
    v = decision.verdict_summary
    cap = decision.capability
    lines = [
        "【자동 판정 요약 — AIAG-VDA 순서】",
        f"  [1] 공정상태: {v.process_stability}",
        f"  [2] 정규성: {v.normality_verdict}",
        f"  [3] Primary KPI: {v.primary_kpi}",
        f"  [4] Cp/Cpk 유효성: {v.cp_cpk_validity}",
        f"  [5] 관리도 이상 패턴: {v.western_electric_summary}",
        f"  [6] 권고 조치: {v.priority_action}",
        f"  공정 레벨: {v.process_level}",
        f"  Subgroup: {v.subgroup_rationality}",
        f"  공정능력 판정: {v.capability_verdict}",
        f"  관리용 관리도 적용: {v.control_chart_deploy}",
    ]
    if cap and cap.cpk_ppk_gap is not None:
        lines.append(f"  Cpk−Ppk Gap: {cap.cpk_ppk_gap:.3f} ({cap.gap_interpretation})")
    if decision.normality.applied_action:
        lines.append(f"  정규성 조치: {decision.normality.applied_action}")
    lines.extend([
        "",
        "【회사 기준 판정 근거 (Decision Log)】",
    ])
    for entry in decision.control_chart.decision_log:
        lines.append(f"  · [{entry.rule_id}] {entry.message}")
    if decision.control_chart.western_electric_violations:
        lines.append("")
        lines.append("【Western Electric Rules】")
        for we_v in decision.control_chart.western_electric_violations:
            loc = ", ".join(str(p) for p in we_v.affected_subgroups)
            lines.append(
                f"  · {we_v.rule_id}: {we_v.rule_name} — {we_v.occurrence_count}회 (subgroup {loc})"
            )
    lines.extend([
        "",
        "【SPC 전문가 해석】",
        f"▶ 경영진 요약: {decision.expert_commentary.executive_summary}",
        f"▶ 실무 해석: {decision.expert_commentary.field_operator_comment}",
        f"▶ 정규성: {decision.expert_commentary.normality_comment}",
        f"▶ 공정능력: {decision.expert_commentary.capability_comment}",
        f"▶ 후속조치: {decision.expert_commentary.followup_action_comment}",
        "",
        "【AIAG-VDA 확장 점검】",
        f"  Pre-control: {decision.aiag_vda_extensions.pre_control_recommendation}",
        f"  기계성능(Cm/Cmk) 필요: {'예' if decision.aiag_vda_extensions.machine_capability_needed else '아니오'}",
        f"  {decision.aiag_vda_extensions.machine_capability.message}",
        f"  Pp/Ppk 기준: {decision.aiag_vda_extensions.pp_ppk_basis_note}",
        f"  Cp/Cpk 기준: {decision.aiag_vda_extensions.cp_cpk_basis_note}",
        "",
        "【리포트 완전성】",
    ])
    rc = decision.aiag_vda_extensions.report_completeness
    status = "완전" if rc.completeness_ok else "누락 있음"
    lines.append(f"  상태: {status}")
    if rc.missing_items:
        lines.append(f"  누락: {', '.join(rc.missing_items)}")
    for w in rc.warnings:
        lines.append(f"  ⚠ {w}")
    return "\n".join(lines)


class ComprehensiveReportGenerator:
    """종합시트(Excel+PDF) + 세부 Excel 시트."""

    def __init__(self, output_dir: str | Path):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate(
        self,
        result: SpcAnalysisResult,
        *,
        charts: ChartPaths,
        raw_sample: pd.DataFrame,
        study_info: dict,
        report_title: str = "SPC 및 공정능력 연구 보고서",
        decision: SpcDecisionResult | None = None,
        file_tag: str | None = None,
    ) -> dict[str, Path]:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        suffix = f"_{safe_tag}" if (safe_tag := _file_tag_slug(file_tag)) else ""
        excel_path = self.output_dir / f"SPC_종합보고서{suffix}_{ts}.xlsx"
        pdf_path = self.output_dir / f"SPC_종합보고서{suffix}_{ts}.pdf"

        self._write_excel(excel_path, result, charts, raw_sample, study_info, report_title, decision)
        self._write_pdf(pdf_path, result, charts, study_info, report_title, raw_sample, decision)

        logger.info("종합 보고서: %s, %s", excel_path, pdf_path)
        return {"excel": excel_path, "pdf": pdf_path, "excel_detail": excel_path}

    def generate_bytes(
        self,
        result: SpcAnalysisResult,
        *,
        charts: ChartPaths,
        raw_sample: pd.DataFrame,
        study_info: dict,
        report_title: str = "SPC 및 공정능력 연구 보고서",
        decision: SpcDecisionResult | None = None,
        file_tag: str | None = None,
    ) -> tuple[bytes, bytes, str]:
        """Excel/PDF 바이트 생성 (디스크 output 저장 없음)."""
        import tempfile

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        suffix = f"_{safe_tag}" if (safe_tag := _file_tag_slug(file_tag)) else ""
        stem = f"SPC_종합보고서{suffix}_{ts}"

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_xlsx:
            excel_path = Path(tmp_xlsx.name)
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp_pdf:
            pdf_path = Path(tmp_pdf.name)
        try:
            self._write_excel(excel_path, result, charts, raw_sample, study_info, report_title, decision)
            excel_bytes = excel_path.read_bytes()
            self._write_pdf(pdf_path, result, charts, study_info, report_title, raw_sample, decision)
            pdf_bytes = pdf_path.read_bytes()
        finally:
            excel_path.unlink(missing_ok=True)
            pdf_path.unlink(missing_ok=True)

        return excel_bytes, pdf_bytes, stem

    def _write_excel(
        self,
        path: Path,
        result: SpcAnalysisResult,
        charts: ChartPaths,
        raw_sample: pd.DataFrame,
        study_info: dict,
        title: str,
        decision: SpcDecisionResult | None = None,
    ) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "종합"

        # 제목
        ws.merge_cells("A1:L1")
        c = ws["A1"]
        c.value = title
        c.font = Font(bold=True, size=14, color="1F4E79")
        c.alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:L2")
        ws["A2"].value = "AIAG / VDA SPC Harmonized Standard — Process Capability Study"
        ws["A2"].font = Font(size=9, italic=True, color="666666")
        ws["A2"].alignment = Alignment(horizontal="center")

        row = 4
        if decision:
            row = write_verdict_strip_on_summary(ws, decision, start_row=3)

        meta_start = row

        # 메타데이터 (좌측)
        meta_rows = self._build_meta_rows(result, study_info, decision)
        for label, value in meta_rows:
            ws.cell(row, 1, label).font = LABEL_FONT
            ws.cell(row, 1).border = BORDER
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
            val_cell = ws.cell(row, 2, value)
            val_cell.font = VALUE_FONT
            val_cell.border = BORDER
            row += 1

        # 공정능력 지표 (우측) — 메타와 동일 시작행 (판정 스트립과 겹치지 않게)
        cap_start = meta_start
        cap_headers = ["지표", "값"]
        for i, h in enumerate(cap_headers):
            cell = ws.cell(cap_start, 6 + i, h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = BORDER

        cap_rows = self._capability_rows(result)
        for i, (k, v) in enumerate(cap_rows):
            r = cap_start + 1 + i
            ws.cell(r, 6, k).font = LABEL_FONT
            ws.cell(r, 6).border = BORDER
            ws.cell(r, 7, v).font = VALUE_FONT
            ws.cell(r, 7).border = BORDER

        # 차트 2x2 배치
        chart_row = row + 1
        positions = [
            (charts.histogram, f"A{chart_row}"),
            (charts.raw_chart, f"G{chart_row}"),
            (charts.prob_plot, f"A{chart_row + 18}"),
            (charts.control_chart, f"G{chart_row + 18}"),
        ]
        for img_path, anchor in positions:
            if img_path.exists():
                img = XLImage(str(img_path))
                img.width = 340
                img.height = 255
                ws.add_image(img, anchor)

        # §20 차트별 전문가 결론·권고
        concl_row = chart_row + 36
        ws.merge_cells(f"A{concl_row}:L{concl_row}")
        ws[f"A{concl_row}"].value = "20. Conclusions / Recommendations (차트별 품질 전문가 점검)"
        ws[f"A{concl_row}"].font = LABEL_FONT
        concl_body_row = concl_row + 1
        concl_end_row = concl_row + 32
        ws.merge_cells(f"A{concl_body_row}:L{concl_end_row}")
        concl_cell = ws[f"A{concl_body_row}"]
        concl_cell.value = _conclusion_text(result, raw_sample, decision)
        concl_cell.alignment = Alignment(wrap_text=True, vertical="top")
        concl_cell.font = Font(size=8)

        for col in range(1, 13):
            ws.column_dimensions[get_column_letter(col)].width = 14

        sheet_registry = self._add_detail_sheets(wb, result, raw_sample, decision)
        if decision:
            add_decision_dashboard_sheet(wb, decision, study_info=study_info, insert_at=0)
            apply_decision_sheet_styles(wb, decision)
        add_glossary_sheet(wb, result)
        add_validation_sheet(wb, result, sheet_registry, decision)
        wb.save(path)
        sanitize_xlsx_formula_file(path)

    def _build_meta_rows(
        self,
        result: SpcAnalysisResult,
        info: dict,
        decision: SpcDecisionResult | None = None,
    ) -> list[tuple[str, str]]:
        cap = result.capability
        rows = [
            ("1. 공정명", info.get("process", "-")),
            ("2. 설비/라인", info.get("machine", "-")),
            ("3. 품목", info.get("item", "-")),
            ("4. 특성(검사항목)", info.get("characteristic", "-")),
            ("5. USL", _fmt_spec_limit(cap.usl) if cap else "-"),
            ("6. LSL", _fmt_spec_limit(cap.lsl) if cap else "-"),
            ("7. 표본수", str(result.normality.n)),
            ("8. Subgroup 크기", str(result.control_limits.subgroup_size or "-")),
            ("9. 관리도 유형", result.control_limits.chart_type),
            ("10. 분석일시", datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("15. 정규성 검정", f"{result.normality.test_name}, p={result.normality.p_value:.4f}"),
            ("16. 정규성 판정", "정규" if result.normality.is_normal else "비정규"),
            ("17. 관리상태", "관리外" if result.out_of_control_points else "관리內"),
        ]
        if info.get("data_source"):
            rows.insert(3, ("데이터 출처", info["data_source"]))
        if info.get("sampling"):
            rows.append(("18. 채취 방식", info["sampling"]))
        rows.append(("19. 역추적", "Excel 시트 「역추적_채취표본」「역추적_Subgroup」 — 분홍=주의"))
        return rows

    def _capability_rows(self, result: SpcAnalysisResult) -> list[tuple[str, str]]:
        if not result.capability:
            return [("공정능력", "USL/LSL 미지정")]
        from scipy import stats as sp_stats

        c = result.capability
        spec = getattr(c, "spec_type", "two_sided")
        pct_above, pct_below = _norm_pct_outside_spec(c, sp_stats)

        rows: list[tuple[str, str]] = [
            ("USL", _fmt_spec_limit(c.usl)),
            ("LSL", _fmt_spec_limit(c.lsl)),
            ("Mean", f"{c.mean:.6f}"),
            ("Std(within)", f"{c.std_within:.6f}"),
            ("Std(overall)", f"{c.std_overall:.6f}"),
        ]
        if spec == "upper_only":
            rows.extend([
                ("Cpk (CWU)", _fmt_cap_number(c.cpk)),
                ("Cpu (CWU)", _fmt_cap_number(c.cpu)),
                ("Ppk", _fmt_cap_number(c.ppk)),
                ("Pp", _fmt_cap_number(c.pp)),
            ])
        elif spec == "lower_only":
            rows.extend([
                ("Cpk (CWL)", _fmt_cap_number(c.cpk)),
                ("Cpl (CWL)", _fmt_cap_number(c.cpl)),
                ("Ppk", _fmt_cap_number(c.ppk)),
                ("Pp", _fmt_cap_number(c.pp)),
            ])
        else:
            rows.extend([
                ("Cp", _fmt_cap_number(c.cp)),
                ("Cpk", _fmt_cap_number(c.cpk)),
                ("Pp", _fmt_cap_number(c.pp)),
                ("Ppk", _fmt_cap_number(c.ppk)),
            ])
        rows.append(("예상 PPM", f"{c.ppm_est:.2f}"))
        if pct_above is not None:
            rows.append(("P% > USL", f"{pct_above:.4f}%"))
        if pct_below is not None:
            rows.append(("P% < LSL", f"{pct_below:.4f}%"))
        return rows

    def _add_detail_sheets(
        self,
        wb: Workbook,
        result: SpcAnalysisResult,
        raw_sample: pd.DataFrame,
        decision: SpcDecisionResult | None = None,
    ) -> dict[str, SheetMeta]:
        sheets: list[tuple[str, pd.DataFrame]] = [
            ("정규성검정", pd.DataFrame([result.normality.to_dict()])),
        ]

        limits_data = []
        cl = result.control_limits
        for chart, lims in [
            ("Xbar", cl.xbar_limits),
            ("S", cl.s_limits),
            ("R", cl.r_limits),
            ("I", cl.i_limits),
            ("MR", cl.mr_limits),
        ]:
            if lims:
                for k, v in lims.items():
                    limits_data.append({"차트": chart, "항목": k, "값": v})
        sheets.append(("관리한계", pd.DataFrame(limits_data)))

        if result.capability:
            sheets.append(("공정능력", pd.DataFrame([result.capability.to_dict()])))

        if result.subgroup_stats is not None:
            sheets.append(("Subgroup통계", result.subgroup_stats))
        if result.individual_stats is not None:
            sheets.append(("Individual통계", result.individual_stats))

        sheets.append(("채취표본", raw_sample))

        for name, df in build_traceability_sheets(raw_sample, result, decision):
            sheets.append((name, df))

        if decision:
            log_rows = [
                {"rule_id": e.rule_id, "message": e.message, "priority": e.priority}
                for e in decision.control_chart.decision_log
            ]
            sheets.append(("판정로그", pd.DataFrame(log_rows)))
            verdict_rows = [{"항목": k, "값": v} for k, v in decision.verdict_summary.to_dict().items()]
            sheets.append(("판정요약", pd.DataFrame(verdict_rows)))
            commentary_rows = [
                {"구분": k, "내용": v}
                for k, v in decision.expert_commentary.to_dict().items()
            ]
            sheets.append(("전문가해석", pd.DataFrame(commentary_rows)))
            aiag_rows = [
                {"항목": "pre_control", "값": decision.aiag_vda_extensions.pre_control_recommendation},
                {"항목": "machine_capability_needed", "값": decision.aiag_vda_extensions.machine_capability_needed},
                {"항목": "machine_capability", "값": decision.aiag_vda_extensions.machine_capability.message},
                {"항목": "completeness_ok", "값": decision.aiag_vda_extensions.report_completeness.completeness_ok},
                {"항목": "missing_items", "값": ", ".join(decision.aiag_vda_extensions.report_completeness.missing_items)},
            ]
            sheets.append(("AIAG_VDA점검", pd.DataFrame(aiag_rows)))

        registry: dict[str, SheetMeta] = {}
        for name, df in sheets:
            if name.startswith("역추적_"):
                registry[name] = self._write_traceability_sheet(wb, name, df)
            else:
                registry[name] = self._write_dataframe_sheet(wb, name, df)
        return registry

    @staticmethod
    def _write_traceability_sheet(wb: Workbook, name: str, df: pd.DataFrame) -> SheetMeta:
        ws = wb.create_sheet(name)
        headers: dict[str, int] = {}
        caution_col: int | None = None
        watch_col: int | None = None
        for c_idx, col in enumerate(df.columns, 1):
            cell = ws.cell(1, c_idx, col)
            cell.fill = TRACE_HEADER_FILL if name.startswith("역추적_") else HEADER_FILL
            cell.font = HEADER_FONT
            headers[str(col)] = c_idx
            if str(col) == "역추적_주의":
                caution_col = c_idx
            if str(col) == "미달/주의":
                watch_col = c_idx

        for r_idx, row in enumerate(df.itertuples(index=False), 2):
            highlight = False
            if caution_col is not None:
                highlight = str(row[caution_col - 1]).strip().upper() == "Y"
            elif watch_col is not None:
                w = str(row[watch_col - 1]).strip().upper()
                highlight = w in ("Y", "△")
            for c_idx, val in enumerate(row, 1):
                if isinstance(val, float) and pd.isna(val):
                    val = None
                cell = ws.cell(r_idx, c_idx, val)
                if highlight:
                    cell.fill = TRACE_CAUTION_FILL

        end_row = max(1, 1 + len(df))
        return SheetMeta(name=name, headers=headers, data_start_row=2, data_end_row=end_row)

    @staticmethod
    def _write_dataframe_sheet(wb: Workbook, name: str, df: pd.DataFrame) -> SheetMeta:
        ws = wb.create_sheet(name)
        headers: dict[str, int] = {}
        for c_idx, col in enumerate(df.columns, 1):
            cell = ws.cell(1, c_idx, col)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            headers[str(col)] = c_idx
        for r_idx, row in enumerate(df.itertuples(index=False), 2):
            for c_idx, val in enumerate(row, 1):
                if isinstance(val, float) and pd.isna(val):
                    val = None
                ws.cell(r_idx, c_idx, val)
        end_row = max(1, 1 + len(df))
        return SheetMeta(name=name, headers=headers, data_start_row=2, data_end_row=end_row)

    def _write_pdf(
        self,
        path: Path,
        result: SpcAnalysisResult,
        charts: ChartPaths,
        study_info: dict,
        title: str,
        raw_sample: pd.DataFrame | None = None,
        decision: SpcDecisionResult | None = None,
    ) -> None:
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
            from reportlab.lib.units import mm
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
        except ImportError as exc:
            raise ImportError("PDF 생성에 reportlab 필요: pip install reportlab") from exc

        font_path = Path(r"C:\Windows\Fonts\malgun.ttf")
        font_name = "Helvetica"
        if font_path.exists():
            pdfmetrics.registerFont(TTFont("MalgunGothic", str(font_path)))
            font_name = "MalgunGothic"

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("Title", parent=styles["Heading1"], fontName=font_name, fontSize=14)
        body_style = ParagraphStyle("Body", parent=styles["Normal"], fontName=font_name, fontSize=9)
        label_style = ParagraphStyle("Label", parent=styles["Normal"], fontName=font_name, fontSize=8)

        doc = SimpleDocTemplate(str(path), pagesize=landscape(A4), leftMargin=12 * mm, rightMargin=12 * mm)
        story = []

        story.append(Paragraph(title, title_style))
        story.append(Paragraph("AIAG / VDA SPC Process Capability Study Report", body_style))
        story.append(Spacer(1, 6))

        meta = self._build_meta_rows(result, study_info, decision)
        meta_table = Table([[a, b] for a, b in meta[:10]], colWidths=[55 * mm, 80 * mm])
        meta_table.setStyle(TableStyle([
            ("FONT", (0, 0), (-1, -1), font_name, 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#E8EEF4")),
        ]))
        story.append(meta_table)
        story.append(Spacer(1, 8))

        cap = self._capability_rows(result)
        cap_table = Table([["지표", "값"]] + list(cap), colWidths=[40 * mm, 35 * mm])
        cap_table.setStyle(TableStyle([
            ("FONT", (0, 0), (-1, -1), font_name, 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ]))
        story.append(cap_table)
        story.append(Spacer(1, 8))

        img_w, img_h = 120 * mm, 55 * mm
        for chart_path in [charts.histogram, charts.raw_chart, charts.prob_plot, charts.control_chart]:
            if chart_path.exists():
                story.append(Image(str(chart_path), width=img_w, height=img_h))
                story.append(Spacer(1, 4))

        concl_style = ParagraphStyle(
            "Conclusion",
            parent=label_style,
            fontSize=7,
            leading=9,
            spaceBefore=4,
        )
        concl_html = _conclusion_text(result, raw_sample, decision).replace("\n", "<br/>")
        story.append(Paragraph("<b>20. Conclusions / Recommendations</b>", label_style))
        story.append(Paragraph(concl_html, concl_style))
        doc.build(story)


def _file_tag_slug(tag: str | None) -> str:
    if not tag or not str(tag).strip():
        return ""
    return safe_filename_slug(str(tag))
