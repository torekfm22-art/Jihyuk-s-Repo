"""Excel 보고서 — 산출값 검증·수식 연계 시트 (영문 Excel 수식)."""
from __future__ import annotations

import re
import shutil
import tempfile
import zipfile
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.spc.constants import A2, A3, C4, D2, I_MR_D2
from src.spc.data_extractor import COLUMN_ALIASES, _col_key
from src.spc.statistics import SpcAnalysisResult

THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
SECTION_FILL = PatternFill("solid", fgColor="D9E2F3")
SECTION_FONT = Font(bold=True, size=10, color="1F4E79")

SHEET_VALIDATION = "검증_수식연계"
PARAM_ROW = 200
SAMPLE_SHEET_NAME = "채취표본"
NORM_SHEET_NAME = "정규성검정"

PVALUE_HEADER_CANDIDATES = ("p_value", "p-value", "pvalue", "P-value", "P_VALUE")

# Excel Open XML "future functions" — _xlfn. 접두사 없으면 Excel 365가 =@STDEV.S 로 표시·저장함
_XLFN_FUNCTIONS: tuple[str, ...] = (
    "STDEV.S",
    "STDEV.P",
    "VAR.S",
    "VAR.P",
    "NORM.S.DIST",
    "NORM.S.INV",
    "CONFIDENCE.NORM",
    "CONFIDENCE.T",
)


def _excel_sheet_prefix(sheet_name: str) -> str:
    """Excel 시트 참조 접두사 (공백·특수문자 있으면 작은따옴표)."""
    if re.search(r"[\s'\[\]!]", sheet_name):
        return f"'{sheet_name}'!"
    return f"{sheet_name}!"


def _alias_keys_for(standard: str) -> set[str]:
    return {_col_key(a) for a in COLUMN_ALIASES.get(standard, ())}


def _header_is_metadata(header: str) -> bool:
    k = _col_key(header)
    for std in COLUMN_ALIASES:
        if std == "value":
            continue
        if k in _alias_keys_for(std):
            return True
    return False


@dataclass
class SheetMeta:
    name: str
    headers: dict[str, int] = field(default_factory=dict)
    data_start_row: int = 2
    data_end_row: int = 1

    def col_letter(self, header: str) -> str | None:
        idx = self.headers.get(header)
        if not idx:
            return None
        return get_column_letter(idx)

    def find_header(
        self,
        *exact: str,
        standard: str | None = None,
    ) -> str | None:
        """헤더명 exact 매칭 → COLUMN_ALIASES 표준명(alias) 매칭."""
        for name in exact:
            if name in self.headers:
                return name
        if standard:
            keys = _alias_keys_for(standard)
            for header in self.headers:
                if _col_key(header) in keys:
                    return header
        return None

    def resolve_measurement_header(self) -> str | None:
        """채취표본 등에서 측정값 열 헤더 자동 탐지."""
        found = self.find_header("value", standard="value")
        if found:
            return found
        for header in self.headers:
            k = _col_key(header)
            if any(tok in k for tok in ("측정", "검사값", "결과", "measure", "weight", "중량")):
                return header
        non_meta = [h for h in self.headers if not _header_is_metadata(h)]
        if len(non_meta) == 1:
            return non_meta[0]
        return None

    def range_ref(self, header: str, *, quote_sheet: bool | None = None) -> str | None:
        col = self.col_letter(header)
        if not col or self.data_end_row < self.data_start_row:
            return None
        s, e = self.data_start_row, self.data_end_row
        if quote_sheet is True or (quote_sheet is None and re.search(r"[\s'\[\]!]", self.name)):
            prefix = f"'{self.name}'!"
        else:
            prefix = _excel_sheet_prefix(self.name)
        return f"{prefix}${col}${s}:${col}${e}"

    def measurement_range_ref(self) -> str | None:
        """측정값 열 전체 데이터 범위 (행·열 자동)."""
        header = self.resolve_measurement_header()
        if not header:
            return None
        return self.range_ref(header, quote_sheet=False)

    def cell_ref(self, header: str, row: int | None = None) -> str | None:
        col = self.col_letter(header)
        if not col:
            return None
        r = self.data_start_row if row is None else row
        return f"{_excel_sheet_prefix(self.name)}${col}${r}"


@dataclass
class ValidationFormulaContext:
    """검증 시트에 쓸 자동 탐지된 Excel 범위·수식."""

    sample_range: str | None = None
    measurement_header: str | None = None
    stdev_expr: str = "-"
    stdev_formula: str = "-"
    norm_pvalue_ref: str = "-"

    @classmethod
    def from_sheets(cls, sheets: dict[str, SheetMeta]) -> ValidationFormulaContext:
        sample = sheets.get(SAMPLE_SHEET_NAME)
        sample_range = sample.measurement_range_ref() if sample else None
        header = sample.resolve_measurement_header() if sample else None
        stdev_expr = f"_xlfn.STDEV.S({sample_range})" if sample_range else "-"
        stdev_formula = to_excel_storage_formula(stdev_expr) if sample_range else "-"

        norm = sheets.get(NORM_SHEET_NAME)
        p_header = None
        if norm:
            p_header = norm.find_header(*PVALUE_HEADER_CANDIDATES)
            if not p_header:
                for h in norm.headers:
                    k = _col_key(h)
                    if k in ("pvalue", "p_value") or (k.startswith("p") and "value" in k):
                        p_header = h
                        break
        norm_ref = norm.cell_ref(p_header, row=2) if norm and p_header else "-"

        return cls(
            sample_range=sample_range,
            measurement_header=header,
            stdev_expr=stdev_expr,
            stdev_formula=stdev_formula,
            norm_pvalue_ref=norm_ref,
        )


def _strip_implicit_at(formula: str) -> str:
    """= @STDEV.S / =@@STDEV.S → =STDEV.S"""
    while formula.startswith("=@"):
        formula = "=" + formula[2:].lstrip()
    if formula.startswith("= @"):
        formula = "=" + formula[3:].lstrip()
    return formula


def _apply_xlfn_prefix(formula_body: str) -> str:
    """Open XML future function 접두사 (_xlfn.) 적용."""
    for fn in sorted(_XLFN_FUNCTIONS, key=len, reverse=True):
        prefixed = f"_xlfn.{fn}"
        if fn in formula_body and prefixed not in formula_body:
            formula_body = formula_body.replace(fn, prefixed)
    return formula_body


def to_excel_storage_formula(expr: str) -> str:
    """
    Excel 저장용 수식 정규화.

    - @ 제거
    - STDEV.S 등 → _xlfn.STDEV.S (Excel 365가 @ 를 붙이지 않도록 Open XML 표준 형식)
    """
    if not expr or expr == "-":
        return "-"
    formula = expr if expr.startswith("=") else f"={expr}"
    formula = _strip_implicit_at(formula)
    body = _apply_xlfn_prefix(formula[1:])
    return f"={body}"


def _fx(expr: str) -> str:
    """셀 기록용 Excel 수식 (=@ 없음, _xlfn. 적용)."""
    return to_excel_storage_formula(expr)


def _set_cell_formula(ws, address: str, formula: str) -> None:
    """셀에 수식 기록."""
    ws[address].value = _fx(formula)


def normalize_workbook_formulas(wb: Workbook) -> None:
    """저장 전 워크북 전체 수식 정규화 (@ 제거, _xlfn. 적용)."""
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith("="):
                    cleaned = _fx(v)
                    if cleaned != v:
                        cell.value = cleaned


def _patch_formula_xml_content(content: str) -> str:
    """<f> 태그 본문: @ 제거 + future function 접두사."""
    content = re.sub(r"^@+", "", content)
    return _apply_xlfn_prefix(content)


def sanitize_xlsx_formula_file(path: str | Path) -> None:
    """
    저장된 xlsx ZIP 내부 시트 XML 수식 패치.

    openpyxl/Excel 365 호환: STDEV.S → _xlfn.STDEV.S, 선행 @ 제거.
    """
    path = Path(path)
    if not path.exists():
        return

    with tempfile.TemporaryDirectory() as tmp:
        tmp_dir = Path(tmp)
        with zipfile.ZipFile(path, "r") as zin:
            zin.extractall(tmp_dir)

        changed = False
        worksheets = tmp_dir / "xl" / "worksheets"
        if worksheets.exists():
            for xml_path in worksheets.glob("*.xml"):
                text = xml_path.read_text(encoding="utf-8")
                original = text

                def repl(match: re.Match[str]) -> str:
                    inner = _patch_formula_xml_content(match.group(1))
                    return f"<f>{inner}</f>"

                text = re.sub(r"<f>([^<]*)</f>", repl, text)
                if text != original:
                    xml_path.write_text(text, encoding="utf-8")
                    changed = True

        if not changed:
            return

        temp_out = path.with_suffix(".xlsx.tmp")
        with zipfile.ZipFile(temp_out, "w", zipfile.ZIP_DEFLATED) as zout:
            for file_path in tmp_dir.rglob("*"):
                if file_path.is_file():
                    arcname = file_path.relative_to(tmp_dir).as_posix()
                    zout.write(file_path, arcname)
        shutil.move(str(temp_out), str(path))


def _q(sheet: str, ref: str) -> str:
    return f"'{sheet}'!{ref}"


def _sample_value_range(ctx: ValidationFormulaContext) -> str | None:
    return ctx.sample_range


def _stdev_sample_formula(ctx: ValidationFormulaContext) -> str:
    return ctx.stdev_expr


def _cell_ref(sheets: dict[str, SheetMeta], sheet_name: str, header: str) -> str:
    meta = sheets.get(sheet_name)
    if not meta:
        return "0"
    resolved = meta.find_header(header) or (header if header in meta.headers else None)
    if not resolved:
        return "0"
    ref = meta.cell_ref(resolved, row=meta.data_start_row)
    return ref if ref else "0"


def _style_header_row(ws, row: int, ncol: int) -> None:
    for c in range(1, ncol + 1):
        cell = ws.cell(row, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _write_section_title(ws, row: int, title: str, ncol: int = 7) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncol)
    c = ws.cell(row, 1, title)
    c.fill = SECTION_FILL
    c.font = SECTION_FONT
    return row + 1


def _validation_table_header(ws, row: int) -> int:
    for i, h in enumerate(
        ["항목", "프로그램 산출", "Excel 검증 수식", "Excel 계산값", "차이", "일치", "연계 시트/범위"],
        1,
    ):
        ws.cell(row, i, h)
    _style_header_row(ws, row, 7)
    return row + 1


def _append_validation_row(
    ws,
    row: int,
    *,
    label: str,
    python_val,
    formula: str,
    link_note: str,
    tol: float = 1e-4,
) -> int:
    ws.cell(row, 1, label).border = BORDER
    b = ws.cell(row, 2, python_val)
    b.border = BORDER
    if isinstance(python_val, (int, float)):
        b.number_format = "0.000000"

    c = ws.cell(row, 3)
    c.value = _fx(formula)
    c.border = BORDER
    c.alignment = Alignment(wrap_text=True)

    d = ws.cell(row, 4)
    if formula not in ("-", ""):
        d.value = f"=C{row}"
        d.number_format = "0.000000"
    else:
        d.value = "-"
    d.border = BORDER

    e = ws.cell(row, 5)
    if isinstance(python_val, (int, float)) and formula not in ("-", ""):
        e.value = _fx(f"ABS(B{row}-D{row})")
        e.number_format = "0.000000"
    else:
        e.value = "-"
    e.border = BORDER

    g = ws.cell(row, 6)
    if isinstance(python_val, (int, float)) and formula not in ("-", ""):
        g.value = _fx(f'IF(ABS(B{row}-D{row})<{tol},"OK","NG")')
    else:
        g.value = "N/A"
    g.border = BORDER

    ws.cell(row, 7, link_note).border = BORDER
    ws.cell(row, 7).alignment = Alignment(wrap_text=True)
    return row + 1


def add_validation_sheet(
    wb: Workbook,
    result: SpcAnalysisResult,
    sheets: dict[str, SheetMeta],
) -> None:
    ctx = ValidationFormulaContext.from_sheets(sheets)
    ws = wb.create_sheet(SHEET_VALIDATION, 2)
    ws["A1"] = "SPC 산출값 검증 · 데이터·수식 연계"
    ws["A1"].font = Font(bold=True, size=12, color="1F4E79")
    ws.merge_cells("A1:G1")
    auto_note = (
        f"측정값 자동 연결: {ctx.sample_range or '(미탐지)'}"
        + (f" [헤더={ctx.measurement_header}]" if ctx.measurement_header else "")
    )
    ws["A2"] = (
        "프로그램 산출값(B열)과 Excel 수식(C열→D열 계산)을 비교합니다. "
        "수식은 영문 Excel 표준 형식(AVERAGE, STDEV.S, MIN 등)이며, "
        "채취표본 시트의 측정값 열·데이터 행 수는 보고서 생성 시 자동 탐지됩니다. "
        f"{auto_note}"
    )
    ws.merge_cells("A2:G2")
    ws["A2"].alignment = Alignment(wrap_text=True)

    row = 4
    row = _write_linkage_map(ws, row, sheets, result, ctx)
    row += 1
    row = _write_sample_stats_block(ws, row, sheets, result, ctx)
    row += 1
    row = _write_control_limits_block(ws, row, sheets, result)
    row += 1
    _write_param_anchor(ws, sheets, result, ctx)
    row += 1
    row = _write_capability_block(ws, row, sheets, result)
    row += 1
    _write_normality_block(ws, row, sheets, result, ctx)
    _pin_fixed_validation_cells(ws, ctx)
    normalize_workbook_formulas(wb)

    for col, w in zip("ABCDEFG", [22, 16, 44, 14, 10, 8, 28], strict=True):
        ws.column_dimensions[col].width = w


def _pin_fixed_validation_cells(ws, ctx: ValidationFormulaContext) -> None:
    """고정 셀(C17, B204, C44) 수식 — 보고서 레이아웃 기준 (=STDEV.S, @ 없음)."""
    if ctx.stdev_formula != "-":
        stdev = _fx(ctx.stdev_formula)
        ws["C17"].value = stdev
        ws["B204"].value = stdev
    d17 = ws["D17"].value
    if d17 is None or d17 == "-" or (isinstance(d17, str) and d17.startswith("=C")):
        ws["D17"] = "=C17"

    if ctx.norm_pvalue_ref != "-":
        ws["C44"] = _fx(ctx.norm_pvalue_ref)
    d44 = ws["D44"].value
    if d44 is None or d44 == "-" or (isinstance(d44, str) and d44.startswith("=C")):
        ws["D44"] = "=C44"
    note = (
        "Excel 기본 함수에 Shapiro-Wilk 없음 → 정규성검정 시트 p_value(C2) 참조. "
        "동일 p-value는 프로그램( scipy ) 산출값입니다."
    )
    if not ws["G44"].value or "Shapiro" not in str(ws["G44"].value):
        ws["G44"] = note


def _write_linkage_map(
    ws, row: int, sheets: dict[str, SheetMeta], result: SpcAnalysisResult, ctx: ValidationFormulaContext
) -> int:
    row = _write_section_title(ws, row, "1. 보고서 데이터 연계 맵")
    for i, h in enumerate(["종합/요약 항목", "출처 시트", "Excel 참조", "설명"], 1):
        ws.cell(row, i, h)
    _style_header_row(ws, row, 4)
    row += 1

    sample_rng = ctx.sample_range or "-"
    hdr_note = ctx.measurement_header or "value"
    cap = "공정능력" if result.capability else "-"
    for a, b, c, d in [
        ("표본·평균·σ_overall", SAMPLE_SHEET_NAME, sample_rng, f"측정값 열 자동({hdr_note})"),
        ("Subgroup Xbar / R / S", "Subgroup통계", "Xbar·R 또는 S", "X-bar 관리도"),
        ("I / MR", "Individual통계", "I·MR 열", "I-MR 관리도"),
        ("관리한계", "관리한계", "차트·항목별 행", "AIAG 상수"),
        ("Cp/Cpk/Pp/Ppk", cap, f"행{PARAM_ROW} 기준값·수식", "본 시트 §4"),
        ("정규성", NORM_SHEET_NAME, ctx.norm_pvalue_ref, "p-value 열 자동 탐지"),
    ]:
        for i, val in enumerate([a, b, c, d], 1):
            ws.cell(row, i, val).border = BORDER
        row += 1
    return row


def _write_sample_stats_block(
    ws, row: int, sheets: dict[str, SheetMeta], result: SpcAnalysisResult, ctx: ValidationFormulaContext
) -> int:
    row = _write_section_title(ws, row, "2. 채취표본 기반 통계 (Excel 재계산)")
    row = _validation_table_header(ws, row)

    sample = sheets.get(SAMPLE_SHEET_NAME)
    rng = _sample_value_range(ctx)
    if not sample or not rng:
        ws.cell(row, 1, "(채취표본 또는 측정값 열 없음 — MES alias·value·측정값 등 확인)")
        return row + 1

    cap = result.capability
    items = [
        ("표본수 n", result.normality.n, f"COUNT({rng})", rng),
        ("평균 Mean", cap.mean if cap else None, f"AVERAGE({rng})", rng),
        ("σ_overall STDEV.S", cap.std_overall if cap else None, _stdev_sample_formula(ctx), rng),
    ]
    for label, py, fn, note in items:
        row = _append_validation_row(ws, row, label=label, python_val=py, formula=fn, link_note=note)
    return row


def _write_control_limits_block(ws, row: int, sheets: dict[str, SheetMeta], result: SpcAnalysisResult) -> int:
    row = _write_section_title(ws, row, "3. 관리도 한계 검증")
    row = _validation_table_header(ws, row)

    cl = result.control_limits
    n = cl.subgroup_size or 0
    chart = result.chart_type

    if chart in ("xbar_r", "xbar_s") and sheets.get("Subgroup통계"):
        sg = sheets["Subgroup통계"]
        xbar_rng = sg.range_ref("Xbar")
        rows: list[tuple] = [
            ("X̿ (Xbar-bar)", cl.center_line, f"AVERAGE({xbar_rng})", xbar_rng),
        ]
        if chart == "xbar_r" and "R" in sg.headers:
            r_rng = sg.range_ref("R")
            rows.append(("R̄", cl.r_limits["CL"] if cl.r_limits else None, f"AVERAGE({r_rng})", r_rng))
            if n in D2:
                rows.append(
                    ("σ_within=R̄/d2", cl.sigma_estimate, f"AVERAGE({r_rng})/{D2[n]}", f"d2={D2[n]}"),
                )
            if cl.xbar_limits and n in A2:
                rows.append(
                    (
                        "Xbar UCL",
                        cl.xbar_limits["UCL"],
                        f"AVERAGE({xbar_rng})+{A2[n]}*AVERAGE({r_rng})",
                        f"A2={A2[n]}",
                    ),
                )
        if chart == "xbar_s" and "S" in sg.headers:
            s_rng = sg.range_ref("S")
            rows.append(("S̄", cl.s_limits["CL"] if cl.s_limits else None, f"AVERAGE({s_rng})", s_rng))
            if n in C4:
                rows.append(
                    ("σ_within=S̄/c4", cl.sigma_estimate, f"AVERAGE({s_rng})/{C4[n]}", f"c4={C4[n]}"),
                )
            if cl.xbar_limits and n in A3:
                rows.append(
                    (
                        "Xbar UCL",
                        cl.xbar_limits["UCL"],
                        f"AVERAGE({xbar_rng})+{A3[n]}*AVERAGE({s_rng})",
                        f"A3={A3[n]}",
                    ),
                )
        for label, py, fn, note in rows:
            row = _append_validation_row(ws, row, label=label, python_val=py, formula=fn, link_note=str(note), tol=1e-5)

    elif chart == "imr" and sheets.get("Individual통계"):
        ind = sheets["Individual통계"]
        i_rng, mr_rng = ind.range_ref("I"), ind.range_ref("MR")
        row = _append_validation_row(
            ws, row, label="Ī", python_val=cl.center_line, formula=f"AVERAGE({i_rng})", link_note=i_rng, tol=1e-5
        )
        if mr_rng and cl.mr_limits:
            row = _append_validation_row(
                ws, row, label="MR̄", python_val=cl.mr_limits["CL"], formula=f"AVERAGE({mr_rng})", link_note=mr_rng, tol=1e-5
            )
            row = _append_validation_row(
                ws,
                row,
                label="σ_within",
                python_val=cl.sigma_estimate,
                formula=f"AVERAGE({mr_rng})/{I_MR_D2}",
                link_note="I-MR d2",
                tol=1e-5,
            )
    else:
        ws.cell(row, 1, "(검증 데이터 없음)")
        row += 1
    return row


def _write_param_anchor(
    ws, sheets: dict[str, SheetMeta], result: SpcAnalysisResult, ctx: ValidationFormulaContext
) -> None:
    """공정능력 수식이 참조하는 기준 셀 (행 PARAM_ROW)."""
    if not result.capability:
        return
    r = PARAM_ROW
    ws.cell(r - 1, 1, "※ 공정능력 수식 기준값 (자동 채움, 수정 가능)").font = Font(italic=True, size=9)
    sample_rng = _sample_value_range(ctx)
    mean_f = (
        f"AVERAGE({sample_rng})"
        if sample_rng
        else _cell_ref(sheets, "공정능력", "평균(Xbar)")
    )
    entries = [
        ("USL", _cell_ref(sheets, "공정능력", "USL")),
        ("LSL", _cell_ref(sheets, "공정능력", "LSL")),
        ("Mean", mean_f),
        ("σ_within", _sigma_within_formula(result, sheets)),
        ("σ_overall", _stdev_sample_formula(ctx)),
    ]
    for i, (lab, ref) in enumerate(entries):
        ws.cell(r + i, 1, lab).font = Font(bold=True, size=9)
        c = ws.cell(r + i, 2)
        if lab == "σ_overall" and ctx.stdev_formula != "-":
            c.value = _fx(ctx.stdev_formula)
        else:
            c.value = _fx(ref)
        c.number_format = "0.000000"


def _sigma_within_formula(result: SpcAnalysisResult, sheets: dict[str, SheetMeta]) -> str:
    cl = result.control_limits
    n = cl.subgroup_size or 0
    chart = result.chart_type
    if chart == "xbar_r" and sheets.get("Subgroup통계"):
        r_rng = sheets["Subgroup통계"].range_ref("R")
        if r_rng and n in D2:
            return f"AVERAGE({r_rng})/{D2[n]}"
    if chart == "xbar_s" and sheets.get("Subgroup통계"):
        s_rng = sheets["Subgroup통계"].range_ref("S")
        if s_rng and n in C4:
            return f"AVERAGE({s_rng})/{C4[n]}"
    if chart == "imr" and sheets.get("Individual통계"):
        mr_rng = sheets["Individual통계"].range_ref("MR")
        if mr_rng:
            return f"AVERAGE({mr_rng})/{I_MR_D2}"
    return _cell_ref(sheets, "공정능력", "σ_within")


def _safe_div(numer: str, denom: str) -> str:
    """Avoid #DIV/0! when anchor cell is zero (English Excel)."""
    return f"IF(({denom})=0,0,({numer})/({denom}))"


def _write_capability_block(ws, row: int, sheets: dict[str, SheetMeta], result: SpcAnalysisResult) -> int:
    row = _write_section_title(ws, row, "4. 공정능력 지표 검증 (AIAG 수식)")
    row = _validation_table_header(ws, row)

    cap = result.capability
    if not cap:
        ws.cell(row, 1, "USL/LSL 미지정")
        return row + 1

    r0 = PARAM_ROW
    usl, lsl, mean, sw, so = f"$B${r0}", f"$B${r0 + 1}", f"$B${r0 + 2}", f"$B${r0 + 3}", f"$B${r0 + 4}"

    checks = [
        ("Cp", cap.cp, _safe_div(f"({usl}-{lsl})", f"6*{sw}"), "6*sigma_within"),
        ("Cpu", cap.cpu, _safe_div(f"({usl}-{mean})", f"3*{sw}"), "3*sigma_within"),
        ("Cpl", cap.cpl, _safe_div(f"({mean}-{lsl})", f"3*{sw}"), "3*sigma_within"),
        (
            "Cpk",
            cap.cpk,
            f"MIN({_safe_div(f'({usl}-{mean})', f'3*{sw}')},{_safe_div(f'({mean}-{lsl})', f'3*{sw}')})",
            "min(Cpu,Cpl)",
        ),
        ("Pp", cap.pp, _safe_div(f"({usl}-{lsl})", f"6*{so}"), "6*sigma_overall"),
        (
            "Ppk",
            cap.ppk,
            f"MIN({_safe_div(f'({usl}-{mean})', f'3*{so}')},{_safe_div(f'({mean}-{lsl})', f'3*{so}')})",
            "min(Ppu,Ppl)",
        ),
    ]
    for label, py, fn, note in checks:
        sheet_val = _cell_ref(sheets, "공정능력", label)
        row = _append_validation_row(
            ws,
            row,
            label=f"{label} (수식검증)",
            python_val=py,
            formula=fn,
            link_note=f"{note}; sheet={sheet_val}",
            tol=0.0002,
        )
        row = _append_validation_row(
            ws,
            row,
            label=f"{label} (시트교차)",
            python_val=py,
            formula=sheet_val,
            link_note="공정능력 row 2",
            tol=0.0002,
        )
    return row


def _write_normality_block(
    ws, row: int, sheets: dict[str, SheetMeta], result: SpcAnalysisResult, ctx: ValidationFormulaContext
) -> int:
    row = _write_section_title(ws, row, "5. 정규성·시트 교차참조")
    row = _validation_table_header(ws, row)

    norm = result.normality
    p_ref = ctx.norm_pvalue_ref if NORM_SHEET_NAME in sheets else "-"
    row = _append_validation_row(
        ws,
        row,
        label="p-value",
        python_val=norm.p_value,
        formula=p_ref,
        link_note=(
            f"{p_ref} 자동 참조. Excel 표준 함수로 Shapiro-Wilk p-value 계산 불가 "
            "(SWTEST/SHAPIRO 없음, Real Statistics 등 애드인만 별도 지원)."
        ),
        tol=1e-6,
    )
    sample_rng = _sample_value_range(ctx)
    n_formula = f"COUNT({sample_rng})" if sample_rng else "-"
    row = _append_validation_row(
        ws, row, label="표본수 n", python_val=norm.n, formula=n_formula, link_note="채취표본"
    )

    if sheets.get("관리한계"):
        for chart_key, lim_key in [("Xbar", "UCL"), ("Xbar", "CL")]:
            py_val = None
            cl = result.control_limits
            if cl.xbar_limits and lim_key in cl.xbar_limits:
                py_val = cl.xbar_limits[lim_key]
            if py_val is not None:
                ref = _lookup_limits_cell(sheets["관리한계"], chart_key, lim_key)
                row = _append_validation_row(
                    ws,
                    row,
                    label=f"관리한계 {chart_key} {lim_key}",
                    python_val=py_val,
                    formula=ref,
                    link_note="관리한계",
                    tol=1e-5,
                )
    return row


def _lookup_limits_cell(meta: SheetMeta, chart: str, item: str) -> str:
    c_chart = meta.col_letter("차트") or "A"
    c_item = meta.col_letter("항목") or "B"
    c_val = meta.col_letter("값") or "C"
    s, e = meta.data_start_row, meta.data_end_row
    return (
        f"SUMIFS('{meta.name}'!${c_val}${s}:${c_val}${e},"
        f"'{meta.name}'!${c_chart}${s}:${c_chart}${e},\"{chart}\","
        f"'{meta.name}'!${c_item}${s}:${c_item}${e},\"{item}\")"
    )
