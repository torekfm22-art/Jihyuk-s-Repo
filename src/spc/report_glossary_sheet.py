"""Excel 보고서 — SPC·공정능력 초보자용 용어 정의 시트."""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from src.spc.statistics import SpcAnalysisResult

THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
SECTION_FILL = PatternFill("solid", fgColor="E2EFDA")
SECTION_FONT = Font(bold=True, size=11, color="375623")
TERM_FONT = Font(bold=True, size=10, color="1F4E79")
FORMULA_FONT = Font(name="Consolas", size=9, color="000000")
TIP_FILL = PatternFill("solid", fgColor="FFF2CC")
FORMULA_FILL = PatternFill("solid", fgColor="F2F2F2")

SHEET_GLOSSARY = "용어_초보가이드"


def add_glossary_sheet(wb: Workbook, result: SpcAnalysisResult | None = None) -> None:
    ws = wb.create_sheet(SHEET_GLOSSARY, 1)

    ws.merge_cells("A1:F1")
    ws["A1"] = "SPC · 공정능력 — 초보자용 용어 설명"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:F2")
    ws["A2"] = (
        "보고서를 처음 보실 때 이 시트부터 읽으시면 됩니다. "
        "§6 산출식은 본 프로그램이 실제로 사용하는 계산식(AIAG/Minitab 기준)입니다. "
        "검증_수식연계 시트에서 Excel로 같은 식을 재현할 수 있습니다."
    )
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A2"].font = Font(size=9, color="444444")

    row = 4
    for section_title, entries in _glossary_sections(result):
        row = _write_section(ws, row, section_title, entries)
        row += 1

    row = _write_formula_section(ws, row, result)
    row += 1

    row = _write_reading_guide(ws, row)
    _write_quick_reference(ws, row + 2)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 18
    ws.freeze_panes = "A4"


def _write_section(ws, row: int, title: str, entries: list[tuple[str, str, str, str]]) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    c = ws.cell(row, 1, title)
    c.fill = SECTION_FILL
    c.font = SECTION_FONT
    c.alignment = Alignment(vertical="center")
    row += 1

    for col, hdr in enumerate(["용어", "한 줄로 이해하기", "조금 더 자세히", "보고서에서 보는 곳"], 1):
        cell = ws.cell(row, col, hdr)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    row += 1

    for term, short, detail, where in entries:
        ws.cell(row, 1, term).font = TERM_FONT
        ws.cell(row, 1).border = BORDER
        for col, text in enumerate([short, detail, where], 2):
            cell = ws.cell(row, col, text)
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.font = Font(size=9)
        row += 1
    return row


def _write_formula_section(ws, row: int, result: SpcAnalysisResult | None) -> int:
    ncol = 5
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncol)
    c = ws.cell(row, 1, "6. 주요 산출값 산출식 (본 프로그램 적용)")
    c.fill = PatternFill("solid", fgColor="FCE4D6")
    c.font = Font(bold=True, size=11, color="833C0C")
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncol)
    ws.cell(
        row,
        1,
        "기호: X̄=평균, σ=표준편차, n=부군 크기, N=총 측정 개수, USL/LSL=규격 상·하한. "
        "상수 A2,A3,d2,c4 등은 부군 크기 n에 따라 AIAG 표에서 자동 적용됩니다.",
    ).font = Font(size=8, italic=True, color="666666")
    ws.cell(row, 1).alignment = Alignment(wrap_text=True)
    row += 2

    for subtitle, entries in _formula_sections(result):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncol)
        ws.cell(row, 1, subtitle).font = Font(bold=True, size=10, color="1F4E79")
        row += 1
        for col, hdr in enumerate(
            ["산출값", "산출식", "변수·기호 설명", "데이터 출처", "검증 시트"],
            1,
        ):
            cell = ws.cell(row, col, hdr)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = BORDER
        row += 1
        for item in entries:
            name, formula, vars_desc, source, verify = item
            ws.cell(row, 1, name).font = TERM_FONT
            ws.cell(row, 1).border = BORDER
            fcell = ws.cell(row, 2, formula)
            fcell.font = FORMULA_FONT
            fcell.fill = FORMULA_FILL
            fcell.border = BORDER
            fcell.alignment = Alignment(wrap_text=True, vertical="top")
            for col, text in enumerate([vars_desc, source, verify], 3):
                cell = ws.cell(row, col, text)
                cell.border = BORDER
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.font = Font(size=9)
            row += 1
        row += 1
    return row


def _formula_sections(
    result: SpcAnalysisResult | None,
) -> list[tuple[str, list[tuple[str, str, str, str, str]]]]:
    chart = result.chart_type if result else "xbar_s"
    n = result.control_limits.subgroup_size if result else 5
    n_txt = f"n={n}" if n else "n=부군크기"

    common: list[tuple[str, list[tuple[str, str, str, str, str]]]] = [
        (
            "6-1. 기초 통계 (모든 분석)",
            [
                (
                    "평균 (Mean, X̄)",
                    "X̄ = (1/N) × Σ xi",
                    "xi = 각 측정값, N = 표본 개수",
                    "채취표본 value 열",
                    "검증 §2 AVERAGE",
                ),
                (
                    "σ_overall",
                    "σ_overall = STDEV.S(xi)  [표본 표준편차, n-1]",
                    "전체 측정값의 산포",
                    "채취표본 value 열",
                    "검증 §2 STDEV.S",
                ),
            ],
        ),
        (
            "6-2. 공정능력 (USL·LSL 입력 시)",
            [
                (
                    "Cp",
                    "Cp = (USL − LSL) / (6 × σ_within)",
                    "규격 폭 대비 공정 산포(위치 무관)",
                    "USL, LSL + σ_within",
                    "검증 §4",
                ),
                (
                    "Cpu",
                    "Cpu = (USL − X̄) / (3 × σ_within)",
                    "상한까지 여유",
                    "동일",
                    "검증 §4",
                ),
                (
                    "Cpl",
                    "Cpl = (X̄ − LSL) / (3 × σ_within)",
                    "하한까지 여유",
                    "동일",
                    "검증 §4",
                ),
                (
                    "Cpk",
                    "Cpk = min(Cpu, Cpl)",
                    "치우친 쪽 기준 실질 능력",
                    "동일",
                    "검증 §4",
                ),
                (
                    "Pp",
                    "Pp = (USL − LSL) / (6 × σ_overall)",
                    "전체 산포 기준 공정 폭",
                    "σ_overall",
                    "검증 §4",
                ),
                (
                    "Ppk",
                    "Ppk = min( (USL−X̄)/(3σ_overall), (X̄−LSL)/(3σ_overall) )",
                    "전체 산포 기준 실질 능력",
                    "동일",
                    "검증 §4",
                ),
                (
                    "예상 PPM",
                    "PPM = [ P(X>USL) + P(X<LSL) ] × 1,000,000\n"
                    "P(X>USL) = 1 − Φ((USL−X̄)/σ_within)\n"
                    "P(X<LSL) = Φ((LSL−X̄)/σ_within)  [Φ=표준정규CDF]",
                    "정규분포 가정",
                    "X̄, σ_within, USL, LSL",
                    "—",
                ),
            ],
        ),
        (
            "6-3. 정규성 검정",
            [
                (
                    "Shapiro-Wilk p-value",
                    "p = Shapiro-Wilk(xi)  (N≤5000)\n"
                    "또는 D'Agostino 검정 (N>5000)",
                    "p ≥ α → 정규, p < α → 비정규",
                    "채취표본 전체 value",
                    "검증 §5 (시트 교차)",
                ),
                (
                    "정규성 판정",
                    "정규  if  p-value ≥ α  (기본 α=0.05)",
                    "α = 유의수준",
                    "정규성검정 시트",
                    "정규성검정",
                ),
            ],
        ),
    ]

    if chart == "xbar_r":
        common.insert(
            1,
            (
                f"6-1b. X-bar R 관리도 ({n_txt})",
                [
                    (
                        "부군 평균 X̄i",
                        "X̄i = (1/n) × Σ xij  (j=1..n, i=부군번호)",
                        "xij = i번째 부군의 j번째 측정값",
                        "원시 데이터 → Subgroup통계",
                        "Subgroup통계 Xbar",
                    ),
                    (
                        "부군 범위 Ri",
                        "Ri = max(xij) − min(xij)",
                        "부군 내 산포",
                        "동일",
                        "Subgroup통계 R",
                    ),
                    (
                        "X̿ (Xbar-bar)",
                        "X̿ = (1/k) × Σ X̄i  (k=부군 개수)",
                        "관리도 중심",
                        "Subgroup통계",
                        "검증 §3",
                    ),
                    (
                        "R̄ (R-bar)",
                        "R̄ = (1/k) × Σ Ri",
                        "평균 범위",
                        "Subgroup통계 R",
                        "검증 §3",
                    ),
                    (
                        "σ_within",
                        f"σ_within = R̄ / d2(n)   [d2({n}) 표 참조]",
                        "단기 공정 산포 추정",
                        "R̄, AIAG 상수 d2",
                        "검증 §3·§4",
                    ),
                    (
                        "Xbar UCL / LCL",
                        f"UCL = X̿ + A2(n)×R̄\nLCL = X̿ − A2(n)×R̄",
                        f"A2({n}) AIAG 상수",
                        "Subgroup통계",
                        "검증 §3",
                    ),
                    (
                        "R 차트 UCL / LCL",
                        f"UCL_R = D4(n)×R̄\nLCL_R = D3(n)×R̄",
                        f"D3,D4({n})",
                        "관리한계 시트",
                        "관리한계",
                    ),
                ],
            ),
        )
    elif chart == "xbar_s":
        common.insert(
            1,
            (
                f"6-1b. X-bar S 관리도 ({n_txt})",
                [
                    (
                        "부군 평균 X̄i",
                        "X̄i = (1/n) × Σ xij",
                        "i=부군 번호",
                        "Subgroup통계 Xbar",
                        "Subgroup통계",
                    ),
                    (
                        "X̿ (Xbar-bar)",
                        "X̿ = (1/k) × Σ X̄i",
                        "관리도 중심",
                        "Subgroup통계",
                        "검증 §3",
                    ),
                    (
                        "부군 표준편차 Si",
                        "Si = STDEV.S(xi1..xin)  [부군 내, n-1]",
                        "부군 내 산포",
                        "Subgroup통계 S",
                        "Subgroup통계",
                    ),
                    (
                        "S̄ (S-bar)",
                        "S̄ = (1/k) × Σ Si",
                        "평균 표준편차",
                        "Subgroup통계",
                        "검증 §3",
                    ),
                    (
                        "σ_within",
                        f"σ_within = S̄ / c4(n)   [c4({n}) 표 참조]",
                        "단기 공정 산포",
                        "S̄, AIAG c4",
                        "검증 §3·§4",
                    ),
                    (
                        "Xbar UCL / LCL",
                        f"UCL = X̿ + A3(n)×S̄\nLCL = X̿ − A3(n)×S̄",
                        f"A3({n})",
                        "Subgroup통계",
                        "검증 §3",
                    ),
                    (
                        "S 차트 UCL / LCL",
                        f"UCL_S = B4(n)×S̄\nLCL_S = B3(n)×S̄",
                        f"B3,B4({n})",
                        "관리한계 시트",
                        "관리한계",
                    ),
                ],
            ),
        )
    elif chart == "imr":
        common.insert(
            1,
            (
                "6-1b. I-MR 관리도",
                [
                    (
                        "이동범위 MRi",
                        "MRi = | xi − x(i−1) |  (i=2..N)",
                        "연속 두 점 차이",
                        "Individual통계 MR",
                        "Individual통계",
                    ),
                    (
                        "MR̄",
                        "MR̄ = (1/(N−1)) × Σ MRi",
                        "평균 이동범위",
                        "Individual통계",
                        "검증 §3",
                    ),
                    (
                        "σ_within",
                        f"σ_within = MR̄ / d2  (d2={1.128}, n=2 등가)",
                        "I-MR 공정 산포",
                        "MR̄",
                        "검증 §3·§4",
                    ),
                    (
                        "I 차트 UCL / LCL",
                        "UCL = X̄ + 2.66 × MR̄\nLCL = X̄ − 2.66 × MR̄",
                        "2.66 = 3/d2 (n=2)",
                        "Individual통계 I",
                        "검증 §3",
                    ),
                    (
                        "MR 차트 UCL",
                        f"UCL_MR = D4 × MR̄  (D4={3.267})",
                        "하한 LCL=0",
                        "관리한계",
                        "관리한계",
                    ),
                ],
            ),
        )

    return common


def _write_reading_guide(ws, row: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row, 1, "■ 보고서 시트 읽는 순서 (추천)").font = SECTION_FONT
    row += 1
    steps = [
        ("1", "용어_초보가이드", "지금 보고 계신 시트 — 용어·판단 기준"),
        ("2", "종합", "한 페이지 요약, 차트, 결론"),
        ("3", "검증_수식연계", "숫자가 맞게 계산됐는지 Excel 수식으로 확인"),
        ("4", "채취표본", "분석에 쓴 측정 데이터"),
        ("5", "Subgroup통계 / Individual통계", "관리도용 요약 통계"),
        ("6", "공정능력 · 정규성검정 · 관리한계", "상세 수치표"),
    ]
    for n, name, desc in steps:
        ws.cell(row, 1, n).border = BORDER
        ws.cell(row, 2, name).font = Font(bold=True, size=9)
        ws.cell(row, 2).border = BORDER
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
        ws.cell(row, 3, desc).border = BORDER
        ws.cell(row, 3).alignment = Alignment(wrap_text=True)
        row += 1
    return row


def _write_quick_reference(ws, row: int) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    cell = ws.cell(row, 1, "■ 자주 쓰는 판단 기준 (일반적인 가이드)")
    cell.fill = TIP_FILL
    cell.font = Font(bold=True, size=10)
    row += 1
    tips = [
        "Cpk ≥ 1.33 : 많은 현장에서 「공정이 규격 안에 잘 들어온다」고 보는 기준(회사별 기준은 다를 수 있음)",
        "p-value ≥ 0.05 : 정규성 검정에서 「데이터가 정규분포와 크게 다르지 않다」고 보는 일반적 기준(유의수준 5%)",
        "관리도 점이 UCL/LCL 밖 : 그 시점에 특별한 원인(이상)이 있었을 가능성 — 원인 조사 권장",
        "Cp는 좋은데 Cpk가 낮음 : 산포는 괜찮지만 평균이 규격 중심에서 벗어남 → 위치 조정 필요",
    ]
    for tip in tips:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row, 1, f"· {tip}").alignment = Alignment(wrap_text=True)
        ws.cell(row, 1).font = Font(size=9)
        row += 1


def _glossary_sections(
    result: SpcAnalysisResult | None,
) -> list[tuple[str, list[tuple[str, str, str, str]]]]:
    chart_hint = "Subgroup통계"
    if result:
        if result.chart_type == "imr":
            chart_hint = "Individual통계"
        elif result.chart_type == "xbar_r":
            chart_hint = "Subgroup통계 (Xbar·R)"

    return [
        (
            "1. 기본 개념",
            [
                (
                    "SPC",
                    "공정이 안정적인지·규격을 지키는지 숫자와 차트로 보는 방법",
                    "Statistical Process Control. 측정값을 시간·LOT별로 모아 관리도와 공정능력으로 판단합니다.",
                    "종합 시트 전체",
                ),
                (
                    "USL / LSL",
                    "규격 상한·하한 (이 범위 안이면 합격)",
                    "Upper/Lower Specification Limit. 도면·검사기준에서 정한 허용 범위입니다.",
                    "종합 5·6항, 공정능력 시트",
                ),
                (
                    "Subgroup (부군)",
                    "관리도용으로 묶은 측정값 한 묶음 (예: 연속 5개)",
                    "같은 조건에서 연속 측정한 n개를 1그룹으로 보고 Xbar, R(또는 S)을 계산합니다.",
                    chart_hint,
                ),
                (
                    "σ (시그마)",
                    "데이터가 얼마나 퍼져 있는지 나타내는 산포",
                    "표준편차. 값이 클수록 측정값이 넓게 퍼져 있습니다.",
                    "공정능력 시트, 검증 시트",
                ),
            ],
        ),
        (
            "2. 공정능력 지표 (Cp · Cpk 등)",
            [
                (
                    "Cp",
                    "규격 폭에 비해 산포가 얼마나 좁은지 (공정 폭)",
                    "Cp = (USL−LSL) / (6σ_within). σ_within은 부군 내·부군 간 변동을 반영한 공정 산포입니다. "
                    "평균 위치는 보지 않고 「퍼짐」만 봅니다.",
                    "종합 우측 표, 공정능력 시트",
                ),
                (
                    "Cpk",
                    "규격 중심에 얼마나 잘 맞춰져 있는지 (실질 능력)",
                    "Cpk = min(Cpu, Cpl). 평균이 USL·LSL 중 한쪽에 치우치면 Cpk가 Cp보다 작아집니다. "
                    "현장에서는 Cpk를 가장 많이 봅니다.",
                    "종합 우측 표, 공정능력 시트",
                ),
                (
                    "Cpu / Cpl",
                    "상한·하한 각각에 대한 여유",
                    "Cpu=(USL−평균)/(3σ_within), Cpl=(평균−LSL)/(3σ_within). Cpk는 둘 중 작은 값입니다.",
                    "공정능력 시트, 검증 시트 §4",
                ),
                (
                    "Pp / Ppk",
                    "전체 데이터 산포로 본 능력 (장기·전체 변동)",
                    "σ_overall(표본 전체 표준편차)을 사용합니다. Ppk가 Cpk보다 작으면 "
                    "시간이 지나며 산포가 커졌거나 추가 변동이 있다는 신호일 수 있습니다.",
                    "공정능력 시트",
                ),
                (
                    "σ_within / σ_overall",
                    "단기 공정 산포 vs 전체 산포",
                    "σ_within: 관리도(R,S,MR)로 추정한 공정 산포. σ_overall: 모든 측정값의 표준편차.",
                    "검증 시트 §2·§3",
                ),
                (
                    "PPM (예상)",
                    "규격 밖으로 나갈 것으로 예상되는 불량 비율(백만분율)",
                    "정규분포를 가정해 USL·LSL 밖 확률을 추정한 값입니다.",
                    "종합 표, 공정능력 시트",
                ),
            ],
        ),
        (
            "3. 정규성 검정",
            [
                (
                    "정규성 검정",
                    "데이터가 「종 모양(정규분포)」에 가까운지 통계적으로 확인",
                    "공정능력·PPM 해석은 정규분포 가정에 기대는 경우가 많아, 먼저 확인하는 단계입니다.",
                    "정규성검정 시트, 종합 15·16항",
                ),
                (
                    "p-value (p값)",
                    "「우연이 아니다」의 정도 — 작을수록 정규분포와 다르다",
                    "p가 작으면(예: 0.05 미만) 정규분포라고 보기 어렵다고 판단합니다. "
                    "p가 크면(예: 0.05 이상) 정규와 크게 다르지 않다고 「일단」 받아들이는 경우가 많습니다.",
                    "정규성검정 시트 p_value 열",
                ),
                (
                    "유의수준 α (알파)",
                    "p-value와 비교하는 기준선 (보통 0.05 = 5%)",
                    "p-value < α 이면 「정규가 아니다」로 판정하는 일반적 규칙입니다. "
                    "α=0.05는 「100번 중 5번까지는 우연히 나올 수 있는 차이」를 허용한다는 뜻에 가깝습니다.",
                    "정규성검정 시트 유의수준",
                ),
                (
                    "Shapiro-Wilk",
                    "표본이 정규분포인지 검정하는 대표적인 방법",
                    "본 프로그램은 표본 수에 따라 Shapiro-Wilk 또는 D'Agostino 검정을 사용합니다.",
                    "정규성검정 시트 검정방법",
                ),
                (
                    "정규 / 비정규 판정",
                    "보고서에 표시된 최종 결론",
                    "p-value ≥ α 이면 「정규」, 그렇지 않으면 「비정규」로 표시합니다. "
                    "비정규일 때는 변환·다른 분석·원인 조사를 검토할 수 있습니다.",
                    "종합 16항",
                ),
            ],
        ),
        (
            "4. 관리도",
            [
                (
                    "관리도",
                    "시간·순서에 따른 측정값이 안정적인지 보는 차트",
                    "평균이 갑자기 튀거나 산포가 커지면 「특별 원인」을 찾도록 돕습니다.",
                    "종합 차트, 관리한계 시트",
                ),
                (
                    "CL / UCL / LCL",
                    "중심선·관리 상한·관리 하한",
                    "CL(Center Line): 평균 수준. UCL/LCL: 통계적으로 기대되는 변동 범위 밖이면 이상 신호로 봅니다.",
                    "관리한계 시트, 종합 관리도 차트",
                ),
                (
                    "X-bar (X̄) 차트",
                    "부군별 평균이 안정적인지 확인",
                    "공정 평균이 서서히 변하거나 갑자기 변했는지 봅니다.",
                    chart_hint + ", 종합 차트",
                ),
                (
                    "R / S 차트",
                    "부군 내 산포(범위·표준편차)가 안정적인지 확인",
                    "R=최대−최소, S=표준편차. 산포가 커지면 공정이 불안정해졌을 수 있습니다.",
                    "Subgroup통계, 관리한계",
                ),
                (
                    "I-MR 차트",
                    "개별 측정값·이동범위 차트 (부군 없을 때)",
                    "한 점씩 측정할 때 사용. I=개별값, MR=연속 두 점 차이의 절대값.",
                    "Individual통계, 관리한계",
                ),
                (
                    "관리內 / 관리外",
                    "관리도 점이 한계 안/밖인지 요약",
                    "관리外: UCL 또는 LCL 밖에 점이 있음 → 원인 조사·조치 검토.",
                    "종합 17항",
                ),
            ],
        ),
        (
            "5. 채취·데이터",
            [
                (
                    "채취표본",
                    "이번 분석에 실제로 사용한 측정 데이터",
                    "MES/QMS에서 필터·샘플링한 결과입니다. LOT·일자·교대 정보가 있을 수 있습니다.",
                    "채취표본 시트",
                ),
                (
                    "일자 랜덤 채취",
                    "날짜별로 골고루 뽑아 편향을 줄이는 방식",
                    "특정 하루 데이터만 몰리지 않도록 설계된 샘플링입니다.",
                    "채취표본, 종합 메타",
                ),
                (
                    "검증_수식연계",
                    "프로그램 결과와 Excel 수식이 같은지 확인하는 시트",
                    "OK/NG로 일치 여부를 표시합니다. 수식 셀을 누르면 어떤 데이터를 참조하는지 볼 수 있습니다.",
                    "검증_수식연계 시트",
                ),
            ],
        ),
    ]
