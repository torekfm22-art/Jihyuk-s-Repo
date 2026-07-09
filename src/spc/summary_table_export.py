"""분석 대상별 SPC 판정 요약표 Excel (결론 — 다항목 종합)."""
from __future__ import annotations

import io
import math
from datetime import datetime
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.spc.characteristic_split import format_split_label, safe_filename_slug
from src.spc.chart_violations import collect_dispersion_violation_points
from src.spc.decision_models import SpcDecisionResult
from src.spc.pipeline import SpcPipelineResult
from src.spc.statistics import SpcAnalysisResult, _cap_round

SUMMARY_COLUMNS = [
    "측정항목",
    "LSL",
    "USL",
    "LCL",
    "CL",
    "UCL",
    "유형",
    "판정",
    "Pp",
    "Ppk",
    "Cp",
    "Cpk",
    "비고",
]

CHART_TYPE_LABELS = {
    "xbar_r": "X bar R",
    "xbar_s": "X bar S",
    "imr": "I-MR",
}

THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
STABLE_FILL = PatternFill("solid", fgColor="C6EFCE")
UNSTABLE_FILL = PatternFill("solid", fgColor="FFC7CE")
STABLE_FONT = Font(bold=True, color="006100", size=9)
UNSTABLE_FONT = Font(bold=True, color="9C0006", size=9)
REMARK_WARN_FONT = Font(size=9, color="9C0006")
REMARK_OK_FONT = Font(size=9, color="000000")
CAP_GOOD_FONT = Font(size=9, color="0070C0")
CAP_NA_FONT = Font(size=9, color="808080")


def iter_leaf_pipeline_results(pipe: SpcPipelineResult) -> list[SpcPipelineResult]:
    """배치(중첩 포함) 시 말단 분석 결과 목록; 단일 분석이면 1건."""
    if not pipe.is_batch:
        return [pipe]
    leaves: list[SpcPipelineResult] = []
    for child in pipe.split_results:
        if child.is_batch:
            leaves.extend(child.split_results)
        else:
            leaves.append(child)
    return leaves


def _fmt_limit(val: float | None, *, digits: int = 4) -> str | float | None:
    if val is None:
        return None
    if isinstance(val, float) and (math.isnan(val) or math.isinf(val)):
        return None
    return round(float(val), digits)


def _fmt_cap(val: float | None) -> str:
    if val is None:
        return "—"
    if isinstance(val, float) and (math.isnan(val) or math.isinf(val)):
        return "—"
    rounded = _cap_round(float(val))
    if isinstance(rounded, str):
        return rounded
    return f"{rounded:.2f}"


def _mean_chart_limits(analysis: SpcAnalysisResult) -> tuple[float | None, float | None, float | None]:
    cl = analysis.control_limits
    if analysis.chart_type == "imr" and cl.i_limits:
        lim = cl.i_limits
    elif cl.xbar_limits:
        lim = cl.xbar_limits
    else:
        return cl.lcl, cl.center_line, cl.ucl
    return lim.get("LCL"), lim.get("CL"), lim.get("UCL")


def _compact_point_refs(points: list[int] | set[int]) -> str:
    """연속 subgroup 번호를 #1~7 형식으로 압축."""
    ordered = sorted({int(p) for p in points})
    if not ordered:
        return ""

    segments: list[str] = []
    start = prev = ordered[0]
    for p in ordered[1:]:
        if p == prev + 1:
            prev = p
            continue
        segments.append(f"#{start}~{prev}" if prev > start else f"#{start}")
        start = prev = p
    segments.append(f"#{start}~{prev}" if prev > start else f"#{start}")
    return ", ".join(segments)


def _point_refs(points: list[int] | set[int], *, limit: int | None = None) -> str:
    """이상점 번호 — 연속 구간 압축 (limit=None이면 전체 표기)."""
    ordered = sorted({int(p) for p in points})
    if not ordered:
        return ""
    if limit is not None and len(ordered) > limit:
        return _compact_point_refs(ordered[:limit]) + f" 외 {len(ordered) - limit}건"
    return _compact_point_refs(ordered)


def _dispersion_chart_label(chart_type: str) -> str:
    return {"xbar_r": "R", "xbar_s": "S", "imr": "MR"}.get(chart_type, "R")


def _normality_field(normality, name: str, default=None):
    return getattr(normality, name, default)


def _resolved_transform(
    decision: SpcDecisionResult,
) -> tuple[bool, str | None]:
    """실제 적용·정규성 확보까지 된 변환만 (method: box_cox | johnson_su)."""
    nd = decision.normality
    cap_d = decision.capability

    if _normality_field(nd, "transform_success"):
        method = str(_normality_field(nd, "transform_method") or "").lower()
        if method in ("box_cox", "johnson_su"):
            return True, method

    if cap_d and cap_d.capability_on_transformed:
        method = str(cap_d.normality_transform_method or "").lower()
        if method in ("box_cox", "johnson_su"):
            return True, method

    for att in _normality_field(nd, "transform_attempts") or []:
        if not att.get("success"):
            continue
        if att.get("is_normal_after") is False:
            continue
        method = str(att.get("method") or "").lower()
        if method in ("box_cox", "johnson_su"):
            return True, method

    return False, None


def _build_normality_remark(
    analysis: SpcAnalysisResult,
    decision: SpcDecisionResult | None,
) -> str:
    if analysis.normality.is_normal:
        return "정규 판정 O"

    if decision:
        applied, method = _resolved_transform(decision)
        if applied and method == "box_cox":
            return "정규성 비정규 판정, Box-cox 변환 O"
        if applied and method == "johnson_su":
            return "정규성 비정규 판정, Johnson 변환 O"

    return "정규성 미충족"


def _collect_mean_chart_anomaly_points(
    analysis: SpcAnalysisResult,
    decision: SpcDecisionResult | None,
) -> set[int]:
    """Xbar/I 관리도 이상점 — R 불안정(보류) 여부와 무관하게 전부 수집."""
    anomaly_points: set[int] = set(analysis.out_of_control_points or [])
    if not decision:
        return anomaly_points

    ctrl = decision.control_chart
    for v in ctrl.western_electric_violations:
        anomaly_points.update(int(p) for p in v.affected_subgroups)

    company = ctrl.company_interpretation
    if company:
        for rule in company.detected_rules:
            rid = str(rule.get("ruleId") or rule.get("rule_id") or "")
            if rid in ("CONTROL_LIMIT_OUT", "OSCILLATION", "TREND", "RUN", "SHIFT"):
                raw = rule.get("matched_points") or rule.get("matchedPoints") or []
                anomaly_points.update(int(p) for p in raw)

    for pat in ctrl.detected_patterns:
        if pat.severity in ("critical", "high", "medium"):
            anomaly_points.update(int(p) for p in pat.affected_points)

    return anomaly_points


def _build_dispersion_remark(analysis: SpcAnalysisResult) -> str:
    label = _dispersion_chart_label(analysis.chart_type)
    disp_ooc = sorted(collect_dispersion_violation_points(analysis))
    if disp_ooc:
        return f"{label} 관리도 {_point_refs(disp_ooc)} 관리상한선 초과"
    return f"{label} 관리도 이상 無"


def _count_spec_violations(
    sample_df: pd.DataFrame | None,
    usl: float | None,
    lsl: float | None,
) -> tuple[int, int]:
    if sample_df is None or sample_df.empty or "value" not in sample_df.columns:
        return 0, 0
    vals = pd.to_numeric(sample_df["value"], errors="coerce").dropna()
    above = int((vals > usl).sum()) if usl is not None else 0
    below = int((vals < lsl).sum()) if lsl is not None else 0
    return above, below


def _build_mean_chart_remark(
    analysis: SpcAnalysisResult,
    decision: SpcDecisionResult | None,
    sample_df: pd.DataFrame | None,
) -> str:
    chart_label = "I" if analysis.chart_type == "imr" else "X bar"
    anomaly_points = _collect_mean_chart_anomaly_points(analysis, decision)

    cap = analysis.capability
    usl = cap.usl if cap else None
    lsl = cap.lsl if cap else None
    above_n, below_n = _count_spec_violations(sample_df, usl, lsl)

    spec_notes: list[str] = []
    if below_n >= 3:
        spec_notes.append("규격하한치 초과 실제값 다수 有")
    elif below_n > 0:
        spec_notes.append(f"규격하한 미만 {below_n}건")
    if above_n >= 3:
        spec_notes.append("규격상한 초과 실제값 다수 有")
    elif above_n > 0:
        spec_notes.append(f"규격상한 초과 {above_n}건")

    if anomaly_points:
        msg = f"{chart_label} 관리도 {_point_refs(anomaly_points)} 이상점 발생"
        if spec_notes:
            msg += ", " + ", ".join(spec_notes)
        return msg

    if spec_notes:
        return f"{chart_label} 관리도 " + ", ".join(spec_notes)
    return f"{chart_label} 관리도 이상 無"


def build_summary_remarks(
    analysis: SpcAnalysisResult,
    decision: SpcDecisionResult | None,
    sample_df: pd.DataFrame | None,
) -> str:
    """비고란 — 정규성 + R/S/MR + X bar/I (이미지 양식)."""
    norm_part = _build_normality_remark(analysis, decision)
    disp_part = _build_dispersion_remark(analysis)
    mean_part = _build_mean_chart_remark(analysis, decision, sample_df)
    return f"{norm_part}, {disp_part}, {mean_part}"


def _capability_values(
    analysis: SpcAnalysisResult,
    decision: SpcDecisionResult | None,
) -> tuple[str, str, str, str]:
    cap_d = decision.capability if decision else None
    cap_a = analysis.capability

    def pick(attr: str) -> float | None:
        if cap_d is not None:
            val = getattr(cap_d, attr, None)
            if val is not None and not (isinstance(val, float) and math.isnan(val)):
                return float(val)
        if cap_a is not None:
            val = getattr(cap_a, attr, None)
            if val is not None and not (isinstance(val, float) and math.isnan(val)):
                return float(val)
        return None

    pp, ppk, cp, cpk = pick("pp"), pick("ppk"), pick("cp"), pick("cpk")
    return _fmt_cap(pp), _fmt_cap(ppk), _fmt_cap(cp), _fmt_cap(cpk)


def build_summary_row(
    result: SpcPipelineResult,
    *,
    split_column: str | None = None,
) -> dict[str, Any]:
    """단일 분석 대상 → 요약표 1행."""
    label = format_split_label(result.characteristic or "-", split_column or result.split_column or "")
    analysis = result.analysis
    decision = result.decision

    if analysis is None:
        return {
            "측정항목": label,
            "LSL": None,
            "USL": None,
            "LCL": None,
            "CL": None,
            "UCL": None,
            "유형": None,
            "판정": None,
            "Pp": None,
            "Ppk": None,
            "Cp": None,
            "Cpk": None,
            "비고": "계량형 결과값 X — SPC 분석 불가 (데이터 부족 또는 비계량)",
        }

    cap = analysis.capability
    lsl = _fmt_limit(cap.lsl if cap else None)
    usl = _fmt_limit(cap.usl if cap else None)
    lcl, cl, ucl = _mean_chart_limits(analysis)
    lcl = _fmt_limit(lcl)
    cl = _fmt_limit(cl)
    ucl = _fmt_limit(ucl)

    chart_type = CHART_TYPE_LABELS.get(analysis.chart_type, analysis.chart_type)
    is_stable = bool(decision.control_chart.is_stable) if decision else not analysis.out_of_control_points
    판정 = "안정" if is_stable else "불안정"

    pp_s, ppk_s, cp_s, cpk_s = _capability_values(analysis, decision)
    remarks = build_summary_remarks(analysis, decision, result.sample_df)

    return {
        "측정항목": label,
        "LSL": lsl,
        "USL": usl,
        "LCL": lcl,
        "CL": cl,
        "UCL": ucl,
        "유형": chart_type,
        "판정": 판정,
        "Pp": pp_s,
        "Ppk": ppk_s,
        "Cp": cp_s,
        "Cpk": cpk_s,
        "비고": remarks,
    }


def build_summary_dataframe(pipe: SpcPipelineResult) -> pd.DataFrame:
    split_col = pipe.split_column
    rows = [
        build_summary_row(leaf, split_column=split_col)
        for leaf in iter_leaf_pipeline_results(pipe)
    ]
    return pd.DataFrame(rows, columns=SUMMARY_COLUMNS)


def _cap_cell_font(val: str) -> Font:
    if val == "—":
        return CAP_NA_FONT
    try:
        num = float(val)
        if num >= 1.33:
            return CAP_GOOD_FONT
    except ValueError:
        pass
    return Font(size=9)


def _remark_font(text: str) -> Font:
    warn_tokens = ("비정규", "미충족", "초과", "이상점", "불안정", "다수", "미만")
    if any(t in text for t in warn_tokens):
        return REMARK_WARN_FONT
    return REMARK_OK_FONT


def write_summary_workbook(
    wb: Workbook,
    pipe: SpcPipelineResult,
    *,
    title: str = "SPC 분석 대상별 판정 요약",
    study_info: dict | None = None,
) -> None:
    ws = wb.active
    ws.title = "판정요약"

    ws.merge_cells("A1:M1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=13, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center")

    info = study_info or {}
    meta_parts = [
        info.get("process"),
        info.get("item"),
        info.get("machine"),
        datetime.now().strftime("%Y-%m-%d %H:%M"),
    ]
    meta = " · ".join(str(p) for p in meta_parts if p)
    if meta:
        ws.merge_cells("A2:M2")
        ws["A2"] = meta
        ws["A2"].font = Font(size=9, italic=True, color="666666")
        ws["A2"].alignment = Alignment(horizontal="center")

    header_row = 4
    for col_idx, col_name in enumerate(SUMMARY_COLUMNS, 1):
        cell = ws.cell(header_row, col_idx, col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    df = build_summary_dataframe(pipe)
    for r_idx, row in enumerate(df.itertuples(index=False), header_row + 1):
        for c_idx, val in enumerate(row, 1):
            col_name = SUMMARY_COLUMNS[c_idx - 1]
            cell = ws.cell(r_idx, c_idx, val)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(col_name == "비고"))

            if col_name == "판정":
                if val == "안정":
                    cell.fill = STABLE_FILL
                    cell.font = STABLE_FONT
                elif val == "불안정":
                    cell.fill = UNSTABLE_FILL
                    cell.font = UNSTABLE_FONT
            elif col_name in ("Pp", "Ppk", "Cp", "Cpk") and isinstance(val, str):
                cell.font = _cap_cell_font(val)
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_name == "비고" and isinstance(val, str):
                cell.font = _remark_font(val)
            elif col_name in ("LSL", "USL", "LCL", "CL", "UCL") and isinstance(val, (int, float)):
                cell.number_format = "0.0000"
                cell.alignment = Alignment(horizontal="right", vertical="center")

    widths = [28, 10, 10, 10, 10, 10, 10, 8, 8, 8, 8, 8, 52]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"


def generate_summary_excel_bytes(
    pipe: SpcPipelineResult,
    *,
    study_info: dict | None = None,
    title: str = "SPC 분석 대상별 판정 요약",
    file_tag: str | None = None,
) -> tuple[bytes, str]:
    wb = Workbook()
    write_summary_workbook(wb, pipe, title=title, study_info=study_info)
    buf = io.BytesIO()
    wb.save(buf)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    slug = f"_{safe_filename_slug(file_tag)}" if file_tag else ""
    fname = f"SPC_판정요약표{slug}_{ts}.xlsx"
    return buf.getvalue(), fname
