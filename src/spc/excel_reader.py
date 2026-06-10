"""Excel 파일 형식 감지 및 자동 읽기 (.xlsx / .xls / CSV / 암호화 Excel)."""
from __future__ import annotations

import io
import logging
from enum import Enum
from pathlib import Path

import pandas as pd

logger = logging.getLogger(__name__)

_XLSX_MAGIC = b"PK\x03\x04"
_XLS_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


class ExcelFormat(str, Enum):
    XLSX = "xlsx"
    XLS = "xls"
    ENCRYPTED = "encrypted"
    CSV = "csv"
    UNKNOWN = "unknown"


def detect_file_format(path: Path) -> ExcelFormat:
    path = Path(path)
    suffix = path.suffix.lower()

    if suffix in (".csv", ".txt"):
        return ExcelFormat.CSV

    with open(path, "rb") as f:
        header = f.read(8)

    if header.startswith(_XLSX_MAGIC):
        return ExcelFormat.XLSX

    if header.startswith(_XLS_MAGIC):
        if _is_encrypted_ole(path):
            return ExcelFormat.ENCRYPTED
        return ExcelFormat.XLS

    return ExcelFormat.UNKNOWN


def _is_encrypted_ole(path: Path) -> bool:
    """Microsoft 암호화/DRM Excel 여부 (EncryptedPackage 스트림)."""
    try:
        import olefile
        if not olefile.isOleFile(str(path)):
            return False
        ole = olefile.OleFileIO(str(path))
        streams = ["/".join(s) for s in ole.listdir()]
        ole.close()
        return any("EncryptedPackage" in s for s in streams)
    except ImportError:
        return False
    except Exception:
        return False


def read_excel_auto(
    path: Path,
    sheet_name: str | int = 0,
    password: str | None = None,
) -> pd.DataFrame:
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {path}")

    fmt = detect_file_format(path)
    logger.info("파일 형식 감지: %s → %s", path.name, fmt.value)

    if fmt == ExcelFormat.ENCRYPTED:
        return _read_encrypted(path, sheet_name, password)

    try:
        if fmt == ExcelFormat.CSV:
            return _read_csv(path)
        if fmt == ExcelFormat.XLSX:
            return _read_excel_smart_header(path, sheet_name)
        if fmt == ExcelFormat.XLS:
            return _read_excel_smart_header(path, sheet_name, engine="calamine")
    except Exception as exc:
        if "not a zip file" in str(exc).lower() and _is_encrypted_ole(path):
            return _read_encrypted(path, sheet_name, password)
        raise ValueError(_format_read_error(path, fmt, exc)) from exc

    raise ValueError(
        f"지원하지 않는 파일 형식: {path.name}\n"
        f"지원: xlsx, xls, csv (또는 암호화 Excel + 비밀번호)"
    )


def _read_excel_smart_header(
    path: Path,
    sheet_name: str | int,
    *,
    engine: str = "openpyxl",
) -> pd.DataFrame:
    """상단 빈 행·제목이 있는 시트에서 헤더 행을 자동 찾아 읽기."""
    from src.spc.data_extractor import detect_header_row_index

    preview = pd.read_excel(
        path, sheet_name=sheet_name, header=None, nrows=50, engine=engine
    )
    header_row = detect_header_row_index(preview)
    df = pd.read_excel(path, sheet_name=sheet_name, header=header_row, engine=engine)
    if header_row > 0:
        logger.info("헤더 행 자동 인식: %s (0-based row %d)", path.name, header_row)
    return df


def _read_csv(path: Path) -> pd.DataFrame:
    for enc in ("utf-8-sig", "cp949", "euc-kr", "utf-8"):
        try:
            return pd.read_csv(path, encoding=enc)
        except UnicodeDecodeError:
            continue
    return pd.read_csv(path)


def _read_xls(path: Path, sheet_name: str | int) -> pd.DataFrame:
    try:
        return _read_excel_smart_header(path, sheet_name, engine="calamine")
    except ImportError:
        pass
    except Exception as exc:
        logger.warning("calamine 실패: %s", exc)

    try:
        return _read_excel_smart_header(path, sheet_name, engine="xlrd")
    except ImportError as exc:
        raise ImportError(
            "구형 Excel(.xls) 읽기: pip install python-calamine"
        ) from exc


def _read_encrypted(path: Path, sheet_name: str | int, password: str | None) -> pd.DataFrame:
    try:
        import msoffcrypto
    except ImportError as exc:
        raise ImportError(
            "암호화된 Excel 파일입니다.\n"
            "해결 방법 (택1):\n"
            "  1) Excel에서 열기 → '다른 이름으로 저장' → 일반 xlsx로 저장\n"
            "  2) pip install msoffcrypto-tool 후 GUI에서 파일 비밀번호 입력"
        ) from exc

    if not password:
        raise ValueError(
            f"'{path.name}'은(는) Microsoft 암호화가 적용된 Excel 파일입니다.\n\n"
            "해결 방법:\n"
            "  ① Excel에서 파일을 연 뒤\n"
            "     '다른 이름으로 저장 → Excel 통합 문서(.xlsx)' 로 저장 (암호 해제)\n"
            "  ② 또는 GUI 하단 '파일 비밀번호' 입력 후 다시 실행"
        )

    decrypted = io.BytesIO()
    with open(path, "rb") as f:
        office = msoffcrypto.OfficeFile(f)
        office.load_key(password=password)
        office.decrypt(decrypted)

    decrypted.seek(0)
    from src.spc.data_extractor import detect_header_row_index

    preview = pd.read_excel(decrypted, sheet_name=sheet_name, header=None, nrows=50, engine="openpyxl")
    header_row = detect_header_row_index(preview)
    decrypted.seek(0)
    return pd.read_excel(decrypted, sheet_name=sheet_name, header=header_row, engine="openpyxl")


def _format_read_error(path: Path, fmt: ExcelFormat, exc: Exception) -> str:
    msg = str(exc)
    if "not a zip file" in msg.lower():
        if _is_encrypted_ole(path):
            return (
                f"'{path.name}'은(는) 암호화된 Excel 파일입니다.\n"
                f"Excel에서 '다른 이름으로 저장'으로 일반 xlsx로 변환하거나 비밀번호를 입력하세요."
            )
        return (
            f"'{path.name}' 확장자는 .xlsx 이지만 실제 형식은 구형 Excel(.xls)입니다.\n"
            f"pip install python-calamine 후 다시 시도하세요.\n"
            f"원본 오류: {msg}"
        )
    return f"'{path.name}' 읽기 실패 ({fmt.value}): {msg}"


def decrypt_excel_to_bytes(path: Path, password: str | None = None) -> io.BytesIO:
    """암호화 Excel → 복호화 BytesIO (openpyxl용)."""
    import msoffcrypto

    if not password:
        raise ValueError(
            f"'{path.name}'은(는) 암호화/DRM Excel입니다.\n"
            "Excel에서 '다른 이름으로 저장 → .xlsx' 로 저장하거나 비밀번호를 입력하세요."
        )
    decrypted = io.BytesIO()
    with open(path, "rb") as f:
        office = msoffcrypto.OfficeFile(f)
        office.load_key(password=password)
        office.decrypt(decrypted)
    decrypted.seek(0)
    return decrypted


def load_workbook_auto(path: Path, password: str | None = None):
    """양식/Excel 파일 → openpyxl Workbook (서식·색상 유지)."""
    from openpyxl import load_workbook

    path = Path(path)
    fmt = detect_file_format(path)

    if fmt == ExcelFormat.XLSX:
        return load_workbook(path, data_only=False)

    if fmt == ExcelFormat.ENCRYPTED:
        buf = decrypt_excel_to_bytes(path, password)
        return load_workbook(buf, data_only=False)

    if fmt == ExcelFormat.XLS:
        raise ValueError(
            f"'{path.name}'은(는) 구형 .xls 형식입니다.\n"
            "Excel에서 'Excel 통합 문서(.xlsx)'로 다시 저장해 주세요."
        )

    raise ValueError(f"지원하지 않는 양식 형식: {path.name}")


def list_sheet_names(path: Path, password: str | None = None) -> list[str]:
    fmt = detect_file_format(path)
    if fmt == ExcelFormat.CSV:
        return ["CSV"]
    if fmt == ExcelFormat.ENCRYPTED:
        if not password:
            return ["(암호화됨 — 비밀번호 필요)"]
        buf = decrypt_excel_to_bytes(path, password)
        xl = pd.ExcelFile(buf, engine="openpyxl")
        return xl.sheet_names
    if fmt == ExcelFormat.XLSX:
        xl = pd.ExcelFile(path, engine="openpyxl")
        return xl.sheet_names
    if fmt == ExcelFormat.XLS:
        xl = pd.ExcelFile(path, engine="calamine")
        return xl.sheet_names
    return []
