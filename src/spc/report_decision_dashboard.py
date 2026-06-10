"""Excel 보고서 — 판정·해석 결과 대시보드 UI 시트."""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.spc.decision_models import SpcDecisionResult

SHEET_DASHBOARD = "판정대시보드"

THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=16, color="1F4E79")
SECTION_FONT = Font(bold=True, size=11, color="1F4E79")
BODY_FONT = Font(size=10)
SMALL_FONT = Font(size=9, color="555555")

FILL_OK = PatternFill("solid", fgColor="C6EFCE")
FONT_OK = Font(bold=True, size=14, color="006100")
FILL_BAD = PatternFill("solid", fgColor="FFC7CE")
FONT_BAD = Font(bold=True, size=14, color="9C0006")
FILL_WARN = PatternFill("solid", fgColor="FFEB9C")
FONT_WARN = Font(bold=True, size=14, color="9C6500")
FILL_NEUTRAL = PatternFill("solid", fgColor="D9E2F3")
FONT_NEUTRAL = Font(bold=True, size=14, color="1F4E79")
FILL_BANNER = PatternFill("solid", fgColor="E8EEF4")
FILL_ALT = PatternFill("solid", fgColor="F5F8FC")

VERDICT_KO = {
    "process_stability": "공정상태",
    "normality_verdict": "정규성",
    "capability_verdict": "공정능력",
    "control_chart_deploy": "관리도 적용",
    "priority_action": "우선 조치",
}

COMMENTARY_KO = {
    "executive_summary": "경영진 요약",
    "control_chart_comment": "관리도 해석",
    "normality_comment": "정규성 해석",
    "capability_comment": "공정능력 해석",
    "followup_action_comment": "후속조치 가이드",
    "field_operator_comment": "현장 실무자 해석",
}

COMPLIANCE_KO = {
    "can_deploy_control_chart": "관리용 관리도 적용",
    "requires_recollection": "데이터 재수집 필요",
    "requires_process_improvement": "공정 개선 필요",
    "requires_control_limit_reset": "관리한계 재설정",
    "requires_control_plan_review": "관리계획서 검토",
    "requires_work_instruction_review": "작업표준 검토",
    "requires_containment": "봉쇄 검토",
    "requires_100pct_inspection": "전수검사 검토",
    "requires_customer_exception_review": "고객 예외 검토",
}


def _style_verdict(text: str) -> tuple[PatternFill, Font]:
    if text in ("안정", "정규", "충분", "가능"):
        return FILL_OK, FONT_OK
    if text in ("불안정", "비정규", "부족", "불가", "판정불가"):
        return FILL_BAD, FONT_BAD
    if text in ("경계", "조건부", "예외적 가능"):
        return FILL_WARN, FONT_WARN
    return FILL_NEUTRAL, FONT_NEUTRAL


def _deploy_label(deploy: str) -> str:
    return {
        "possible": "가능",
        "not_possible": "불가",
        "exceptional": "예외적 가능",
        "undetermined": "판정불가",
    }.get(deploy, deploy)


def _set_cell(
    ws: Worksheet,
    row: int,
    col: int,
    value,
    *,
    font: Font | None = None,
    fill: PatternFill | None = None,
    align: Alignment | None = None,
    border: Border | None = BORDER,
) -> None:
    cell = ws.cell(row, col)
    cell.value = value
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if align:
        cell.alignment = align
    if border:
        cell.border = border


def _merge_set(
    ws: Worksheet,
    row: int,
    col_start: int,
    col_end: int,
    value,
    **style,
) -> None:
    """병합 후 좌상단 셀에만 값 기록 (MergedCell read-only 방지)."""
    if col_end > col_start:
        ws.merge_cells(
            start_row=row, start_column=col_start,
            end_row=row, end_column=col_end,
        )
    _set_cell(ws, row, col_start, value, **style)


def add_decision_dashboard_sheet(
    wb: Workbook,
    decision: SpcDecisionResult,
    *,
    study_info: dict | None = None,
    insert_at: int = 0,
) -> None:
    """판정·해석 결과 전용 대시보드 시트 (색상 카드·체크리스트·코멘트)."""
    ws = wb.create_sheet(SHEET_DASHBOARD, insert_at)
    ws.sheet_view.showGridLines = False

    for col, width in enumerate([14, 16, 14, 16, 18, 18], 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    info = study_info or {}
    v = decision.verdict_summary
    comp = decision.compliance

    ws.merge_cells("A1:F1")
    _set_cell(ws, 1, 1, "SPC 자동 판정 · 해석 대시보드", font=TITLE_FONT, border=None)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    meta_line = "  |  ".join(
        x for x in [
            info.get("process"),
            info.get("characteristic"),
            info.get("item"),
            f"단계: {decision.metadata.stage}",
        ] if x and x != "-"
    )
    ws.merge_cells("A2:F2")
    _set_cell(ws, 2, 1, meta_line or "공정능력 연구", font=SMALL_FONT, border=None)
    ws["A2"].alignment = Alignment(horizontal="center")

    # ── KPI 카드 4개 ──
    cards = [
        ("공정상태", v.process_stability),
        ("정규성", v.normality_verdict),
        ("공정능력", v.capability_verdict),
        ("관리도 적용", v.control_chart_deploy),
    ]
    for i, (title, val) in enumerate(cards):
        col = 1 + i
        fill, font = _style_verdict(val)
        _set_cell(ws, 4, col, title, font=Font(bold=True, size=10, color="555555"), fill=FILL_BANNER)
        ws.cell(4, col).alignment = Alignment(horizontal="center")
        _set_cell(ws, 5, col, val, font=font, fill=fill)
        ws.cell(5, col).alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[5].height = 36

    # ── 우선 조치 배너 ──
    ws.merge_cells("A7:F7")
    _set_cell(ws, 7, 1, f"▶ 우선 조치: {v.priority_action}", font=Font(bold=True, size=11, color="1F4E79"), fill=FILL_BANNER)
    ws["A7"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[7].height = 28

    row = 9
    row = _write_section_header(ws, row, "회사 기준 판정 · 후속조치 체크")
    for key, label in COMPLIANCE_KO.items():
        val = getattr(comp, key, None)
        if key == "can_deploy_control_chart":
            display = _deploy_label(str(val))
            status = "●" if val in ("possible", "exceptional") else "○"
        elif isinstance(val, bool):
            display = "필요" if val else "해당 없음"
            status = "●" if val else "—"
        else:
            display = str(val)
            status = "—"
        fill = FILL_WARN if (isinstance(val, bool) and val) or val == "exceptional" else None
        if isinstance(val, bool) and val:
            fill = FILL_WARN
        if key == "can_deploy_control_chart" and val == "not_possible":
            fill = FILL_BAD
        if key == "can_deploy_control_chart" and val == "possible":
            fill = FILL_OK
        _set_cell(ws, row, 1, status, align=Alignment(horizontal="center"))
        _set_cell(ws, row, 2, label, font=BODY_FONT)
        _merge_set(ws, row, 3, 5, display, font=Font(bold=True, size=10), fill=fill)
        row += 1

    if comp.priority_actions:
        _merge_set(
            ws, row, 1, 5,
            "권고 조치: " + " → ".join(comp.priority_actions),
            font=BODY_FONT,
            fill=FILL_BANNER,
        )
        ws.cell(row, 1).alignment = Alignment(wrap_text=True)
        row += 2

    row = _write_section_header(ws, row, "판정 근거 (Decision Log)")
    _set_cell(ws, row, 1, "우선순위", font=HEADER_FONT, fill=HEADER_FILL)
    _set_cell(ws, row, 2, "규칙 ID", font=HEADER_FONT, fill=HEADER_FILL)
    _merge_set(ws, row, 3, 5, "판정 내용", font=HEADER_FONT, fill=HEADER_FILL)
    row += 1
    for i, entry in enumerate(decision.control_chart.decision_log):
        fill = FILL_ALT if i % 2 else None
        _set_cell(ws, row, 1, entry.priority, font=BODY_FONT, fill=fill)
        _set_cell(ws, row, 2, entry.rule_id, font=Font(bold=True, size=9), fill=fill)
        _merge_set(ws, row, 3, 5, entry.message, font=BODY_FONT, fill=fill)
        ws.cell(row, 3).alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    row += 1
    row = _write_section_header(ws, row, "SPC 전문가 해석")
    for key, label in COMMENTARY_KO.items():
        text = getattr(decision.expert_commentary, key, "")
        _merge_set(ws, row, 1, 5, f"■ {label}", font=SECTION_FONT, fill=FILL_BANNER, border=None)
        row += 1
        _merge_set(ws, row, 1, 5, text, font=BODY_FONT, border=None)
        ws.cell(row, 1).alignment = Alignment(wrap_text=True, vertical="top")
        est_lines = max(3, len(text) // 60 + 1)
        ws.row_dimensions[row].height = min(120, 16 * est_lines)
        row += 2

    row = _write_section_header(ws, row, "AIAG-VDA 확장 점검")
    aiag = decision.aiag_vda_extensions
    rc = aiag.report_completeness
    aiag_items = [
        ("Pre-control 권고", aiag.pre_control_recommendation),
        ("기계성능(Cm/Cmk) 필요", "예" if aiag.machine_capability_needed else "아니오"),
        ("기계성능 모듈", aiag.machine_capability.message),
        ("리포트 완전성", "완전" if rc.completeness_ok else "누락 있음"),
        ("누락 항목", ", ".join(rc.missing_items) if rc.missing_items else "없음"),
        ("Pp/Ppk 기준", aiag.pp_ppk_basis_note),
        ("Cp/Cpk 기준", aiag.cp_cpk_basis_note),
    ]
    for label, val in aiag_items:
        _set_cell(ws, row, 1, label, font=Font(bold=True, size=9))
        fill = FILL_BAD if label == "리포트 완전성" and not rc.completeness_ok else None
        if label == "리포트 완전성" and rc.completeness_ok:
            fill = FILL_OK
        _merge_set(ws, row, 2, 5, val, font=BODY_FONT, fill=fill)
        ws.cell(row, 2).alignment = Alignment(wrap_text=True)
        row += 1

    ws.freeze_panes = "A4"


def _write_section_header(ws: Worksheet, row: int, title: str) -> int:
    _merge_set(
        ws, row, 1, 5, title,
        font=SECTION_FONT,
        fill=PatternFill("solid", fgColor="D9E2F3"),
        border=None,
    )
    ws.row_dimensions[row].height = 22
    return row + 1


def write_verdict_strip_on_summary(ws: Worksheet, decision: SpcDecisionResult, start_row: int = 3) -> int:
    """종합 시트 상단에 판정 요약 스트립 삽입. 다음 데이터 시작 행 반환."""
    v = decision.verdict_summary
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=12)
    _set_cell(ws, start_row, 1, "▣ 자동 판정 요약", font=SECTION_FONT, border=None)

    r = start_row + 1
    cards = [
        ("공정", v.process_stability),
        ("정규성", v.normality_verdict),
        ("공정능력", v.capability_verdict),
        ("관리도", v.control_chart_deploy),
    ]
    col_span = 3
    for i, (title, val) in enumerate(cards):
        c0 = 1 + i * col_span
        c1 = c0 + col_span - 1
        fill, font = _style_verdict(val)
        _merge_set(ws, r, c0, c1, f"{title}: {val}", font=font, fill=fill)
        ws.cell(r, c0).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 28

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
    _set_cell(ws, r, 1, f"우선 조치: {v.priority_action}", font=Font(bold=True, size=10), fill=FILL_BANNER)
    ws.cell(r, 1).alignment = Alignment(horizontal="left", wrap_text=True)
    ws.row_dimensions[r].height = 22
    return r + 2


def apply_decision_sheet_styles(wb: Workbook, decision: SpcDecisionResult) -> None:
    """판정요약·전문가해석·판정로그·AIAG 시트 서식 개선."""
    if "판정요약" in wb.sheetnames:
        style_verdict_summary_sheet(wb["판정요약"], decision)
    if "전문가해석" in wb.sheetnames:
        style_commentary_sheet(wb["전문가해석"], decision)
    if "판정로그" in wb.sheetnames:
        style_log_sheet(wb["판정로그"], decision)
    if "AIAG_VDA점검" in wb.sheetnames:
        style_aiag_sheet(wb["AIAG_VDA점검"], decision)


def style_verdict_summary_sheet(ws: Worksheet, decision: SpcDecisionResult) -> None:
    ws.delete_rows(1, ws.max_row)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 40
    _set_cell(ws, 1, 1, "항목", font=HEADER_FONT, fill=HEADER_FILL)
    _set_cell(ws, 1, 2, "판정 결과", font=HEADER_FONT, fill=HEADER_FILL)
    for i, (key, label) in enumerate(VERDICT_KO.items(), 2):
        val = getattr(decision.verdict_summary, key)
        _set_cell(ws, i, 1, label, font=Font(bold=True, size=10))
        fill, font = _style_verdict(val)
        _set_cell(ws, i, 2, val, font=font, fill=fill)
        ws.cell(i, 2).alignment = Alignment(horizontal="center")


def style_commentary_sheet(ws: Worksheet, decision: SpcDecisionResult) -> None:
    ws.delete_rows(1, ws.max_row)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 90
    row = 1
    for key, label in COMMENTARY_KO.items():
        text = getattr(decision.expert_commentary, key, "")
        _merge_set(ws, row, 1, 2, f"■ {label}", font=SECTION_FONT, fill=FILL_BANNER, border=None)
        row += 1
        _merge_set(ws, row, 1, 2, text, font=BODY_FONT, border=None)
        ws.cell(row, 1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = min(100, max(36, len(text) // 50 * 14 + 20))
        row += 2


def style_log_sheet(ws: Worksheet, decision: SpcDecisionResult) -> None:
    ws.delete_rows(1, ws.max_row)
    ws.sheet_view.showGridLines = False
    for col, w in [("A", 10), ("B", 18), ("C", 70), ("D", 10)]:
        ws.column_dimensions[col].width = w
    _set_cell(ws, 1, 1, "우선순위", font=HEADER_FONT, fill=HEADER_FILL)
    _set_cell(ws, 1, 2, "규칙 ID", font=HEADER_FONT, fill=HEADER_FILL)
    _merge_set(ws, 1, 3, 4, "판정 내용", font=HEADER_FONT, fill=HEADER_FILL)
    row = 2
    for i, entry in enumerate(decision.control_chart.decision_log):
        fill = FILL_ALT if i % 2 else None
        _set_cell(ws, row, 1, entry.priority, fill=fill)
        _set_cell(ws, row, 2, entry.rule_id, font=Font(bold=True, size=9), fill=fill)
        _merge_set(ws, row, 3, 4, entry.message, font=BODY_FONT, fill=fill)
        ws.cell(row, 3).alignment = Alignment(wrap_text=True)
        row += 1


def style_aiag_sheet(ws: Worksheet, decision: SpcDecisionResult) -> None:
    ws.delete_rows(1, ws.max_row)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 80
    aiag = decision.aiag_vda_extensions
    rc = aiag.report_completeness
    _set_cell(ws, 1, 1, "점검 항목", font=HEADER_FONT, fill=HEADER_FILL)
    _set_cell(ws, 1, 2, "결과", font=HEADER_FONT, fill=HEADER_FILL)
    items = [
        ("Pre-control 권고", aiag.pre_control_recommendation),
        ("기계성능 필요", "예" if aiag.machine_capability_needed else "아니오"),
        ("Cm/Cmk 모듈", aiag.machine_capability.message),
        ("리포트 완전성", "✓ 완전" if rc.completeness_ok else "✗ 누락"),
        ("누락 항목", ", ".join(rc.missing_items) or "없음"),
    ]
    for i, (label, val) in enumerate(items, 2):
        _set_cell(ws, i, 1, label, font=Font(bold=True, size=10))
        fill = None
        if label == "리포트 완전성":
            fill = FILL_OK if rc.completeness_ok else FILL_BAD
        _set_cell(ws, i, 2, val, font=BODY_FONT, fill=fill)
        ws.cell(i, 2).alignment = Alignment(wrap_text=True)
