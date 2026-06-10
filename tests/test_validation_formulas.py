"""검증_수식연계 시트 수식 형식 테스트."""
from __future__ import annotations

import zipfile
from pathlib import Path

from openpyxl import Workbook

from src.spc.report_validation_sheet import (
    _fx,
    normalize_workbook_formulas,
    sanitize_xlsx_formula_file,
    to_excel_storage_formula,
)


def test_fx_strips_implicit_at():
    assert _fx("=@STDEV.S(A1:A10)") == "=_xlfn.STDEV.S(A1:A10)"
    assert _fx("=@@STDEV.S(A1:A10)") == "=_xlfn.STDEV.S(A1:A10)"
    assert _fx("STDEV.S(A1:A10)") == "=_xlfn.STDEV.S(A1:A10)"
    assert _fx("=STDEV.S(A1:A10)") == "=_xlfn.STDEV.S(A1:A10)"


def test_to_excel_storage_formula_idempotent():
    once = to_excel_storage_formula("=STDEV.S('채취표본'!$M$2:$M$10)")
    twice = to_excel_storage_formula(once)
    assert once == twice == "=_xlfn.STDEV.S('채취표본'!$M$2:$M$10)"


def test_normalize_workbook_strips_at():
    wb = Workbook()
    ws = wb.active
    ws["C17"] = "=@STDEV.S('채취표본'!$M$2:$M$10)"
    ws["B204"] = "=@STDEV.S('채취표본'!$M$2:$M$10)"
    normalize_workbook_formulas(wb)
    assert ws["C17"].value == "=_xlfn.STDEV.S('채취표본'!$M$2:$M$10)"
    assert "@STDEV" not in str(ws["C17"].value)


def test_sanitize_xlsx_xml_patch(tmp_path: Path):
    out = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["C17"] = "=STDEV.S(A1:A10)"
    wb.save(out)
    sanitize_xlsx_formula_file(out)
    with zipfile.ZipFile(out) as z:
        xml = z.read("xl/worksheets/sheet1.xml").decode("utf-8")
    assert "_xlfn.STDEV.S" in xml
    assert "@STDEV" not in xml
